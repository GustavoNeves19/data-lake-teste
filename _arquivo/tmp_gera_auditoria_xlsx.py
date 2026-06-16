"""Gera planilha de auditoria RFV Hospitalar para reunião com Alves/Diego."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

creds = service_account.Credentials.from_service_account_file(
    r'C:\teste\sapient-metrics.json',
    scopes=["https://www.googleapis.com/auth/cloud-platform"]
)
client = bigquery.Client(credentials=creds, project='sapient-metrics-492914-m7')

# === 1. Excel original ===
path_excel = r'C:\Users\gusta\Downloads\RFV Hospitalar 01-04-2025 até 30-04-2026.xlsx'
df_excel = pd.read_excel(path_excel, sheet_name='Base inicial - Geral', usecols=[0, 1, 2])
df_excel.columns = ['cliente', 'data_venda', 'valor']
df_excel = df_excel.dropna(subset=['cliente'])
df_excel['valor'] = pd.to_numeric(df_excel['valor'])
excel_g = df_excel.groupby('cliente').agg(
    pedidos=('data_venda', 'count'),
    valor=('valor', 'sum'),
    ultima=('data_venda', 'max'),
).reset_index()
excel_g['k'] = excel_g['cliente'].str.upper().str.strip()

# === 2. Sistema ===
df_sys = client.query("""
SELECT c.partner_name AS cliente, COUNT(DISTINCT o.order_number) AS pedidos_sys,
       ROUND(SUM(o.total_amount), 2) AS valor_sys
FROM `sapient-metrics-492914-m7.dm_orders.fact_sales_order` o
JOIN `sapient-metrics-492914-m7.silver_comercial.param_com_rfv_carteira` c
  ON c.partner_code = o.partner_code AND c.is_active = TRUE
  AND c.salesperson_name NOT IN ('Eduardo', 'Karina')
JOIN `sapient-metrics-492914-m7.dm_orders.dim_operation_nature` n
  ON n.nature_code = o.nature_code AND n.financial_flag = 'F'
WHERE o.order_status IN (3, 4) AND c.rfv_familia = 'HOSPITALAR'
  AND o.order_date BETWEEN '2025-04-01' AND '2026-04-30'
GROUP BY 1
""").to_dataframe()
df_sys['k'] = df_sys['cliente'].str.upper().str.strip()

# === 3. Faltantes ===
so_excel = excel_g[~excel_g['k'].isin(df_sys['k'])].copy()
faltantes = so_excel[['cliente', 'pedidos', 'valor', 'ultima']].sort_values('valor', ascending=False).reset_index(drop=True)

df_rej = pd.read_csv(r'sql/silver_comercial/carteira_rejeitados.csv')
df_rej['k'] = df_rej['planilha_nome'].str.upper().str.strip()
faltantes['k'] = faltantes['cliente'].str.upper().str.strip()
faltantes = faltantes.merge(df_rej[['k', 'bq_nome', 'score', 'match_type']], on='k', how='left')
faltantes['status_origem'] = faltantes['bq_nome'].apply(
    lambda x: 'REJEITADO (fuzzy match incorreto)' if pd.notna(x) else 'NUNCA PROCESSADO (nao esta na fonte da carteira)'
)
faltantes = faltantes.drop(columns=['k'])

# === 4. Naturezas excluidas ===
df_nat = client.query("""
SELECT
  o.nature_code, ANY_VALUE(n.nature_name) AS nat_name,
  ANY_VALUE(n.financial_flag) AS financial_flag,
  COUNT(DISTINCT c.partner_name) AS clientes,
  COUNT(DISTINCT o.order_number) AS pedidos,
  ROUND(SUM(o.total_amount), 2) AS faturamento
FROM `sapient-metrics-492914-m7.dm_orders.fact_sales_order` o
JOIN `sapient-metrics-492914-m7.silver_comercial.param_com_rfv_carteira` c
  ON c.partner_code = o.partner_code AND c.is_active = TRUE
  AND c.salesperson_name NOT IN ('Eduardo', 'Karina')
LEFT JOIN `sapient-metrics-492914-m7.dm_orders.dim_operation_nature` n
  ON n.nature_code = o.nature_code
WHERE o.order_status IN (3, 4) AND c.rfv_familia = 'HOSPITALAR'
  AND o.order_date BETWEEN '2025-04-01' AND '2026-04-30'
  AND (n.financial_flag != 'F' OR n.financial_flag IS NULL)
GROUP BY 1
ORDER BY faturamento DESC
""").to_dataframe()
df_nat['faturamento'] = pd.to_numeric(df_nat['faturamento'])
df_nat['suspeito'] = df_nat['nat_name'].apply(
    lambda x: 'SIM' if 'VENDA' in str(x) else 'NAO'
)

# === 5. Drill cliente x natureza ===
df_extra = client.query("""
SELECT
  c.partner_name AS cliente, o.nature_code,
  ANY_VALUE(n.nature_name) AS nat_name,
  ANY_VALUE(n.financial_flag) AS flag,
  COUNT(DISTINCT o.order_number) AS pedidos,
  ROUND(SUM(o.total_amount), 2) AS valor_extra
FROM `sapient-metrics-492914-m7.dm_orders.fact_sales_order` o
JOIN `sapient-metrics-492914-m7.silver_comercial.param_com_rfv_carteira` c
  ON c.partner_code = o.partner_code AND c.is_active = TRUE
  AND c.salesperson_name NOT IN ('Eduardo', 'Karina')
LEFT JOIN `sapient-metrics-492914-m7.dm_orders.dim_operation_nature` n
  ON n.nature_code = o.nature_code
WHERE o.order_status IN (3, 4) AND c.rfv_familia = 'HOSPITALAR'
  AND o.order_date BETWEEN '2025-04-01' AND '2026-04-30'
  AND (n.financial_flag != 'F' OR n.financial_flag IS NULL)
GROUP BY 1, 2
ORDER BY valor_extra DESC
LIMIT 200
""").to_dataframe()
df_extra['valor_extra'] = pd.to_numeric(df_extra['valor_extra'])

# === 6. Comparativo segmento ===
seg_data = [
    ('Campeoes', 80, 52),
    ('Fieis', 47, 77),
    ('Fieis em potencial', 112, 73),
    ('Novos clientes', 38, 8),
    ('Promessas', 28, 33),
    ('Precisando de atencao', 6, 5),
    ('Quase dormentes', 63, 77),
    ('Nao pode perder', 11, 11),
    ('Em risco', 21, 38),
    ('Hibernando', 41, 23),
    ('Perdidos', 339, 358),
]
df_seg = pd.DataFrame(seg_data, columns=['Segmento', 'Excel (planilha)', 'Sistema (dashboard)'])
df_seg['Delta (Sistema - Excel)'] = df_seg['Sistema (dashboard)'] - df_seg['Excel (planilha)']

print(f'Faltantes={len(faltantes)} | Naturezas excluidas={len(df_nat)} | Drill={len(df_extra)}')

# ============ EXCEL ============
wb = Workbook()
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FILL_PRIMARY = PatternFill('solid', fgColor='1E1882')
FILL_HEADER = PatternFill('solid', fgColor='4844C8')
FILL_WARN = PatternFill('solid', fgColor='FFE066')
FILL_DANGER = PatternFill('solid', fgColor='F8C9C9')
FILL_OK = PatternFill('solid', fgColor='C9F8D2')
FILL_LIGHT = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_hdr(cell, fill=FILL_HEADER):
    cell.font = ARIAL_HDR
    cell.fill = fill
    cell.alignment = CENTER
    cell.border = BORDER


def style_cell(cell, fmt=None):
    cell.font = ARIAL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


# Aba 1 — Resumo
ws = wb.active
ws.title = 'Resumo Executivo'
ws.merge_cells('A1:E1')
ws['A1'] = 'AUDITORIA RFV HOSPITALAR - Excel (planilha original) x Sistema (Dashboard Nevoni)'
ws['A1'].font = ARIAL_TITLE
ws['A1'].fill = FILL_PRIMARY
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 30

info = [
    ('Periodo auditado:', '01/04/2025 a 30/04/2026'),
    ('Familia RFV:', 'HOSPITALAR'),
    ('Fonte Excel:', 'RFV Hospitalar 01-04-2025 ate 30-04-2026.xlsx'),
    ('Fonte Sistema:', 'silver_comercial.silver_com_rfv_score'),
    ('Gerado em:', datetime.now().strftime('%d/%m/%Y %H:%M')),
]
for i, (k, v) in enumerate(info, start=3):
    ws.cell(i, 1, k).font = ARIAL_BOLD
    ws.cell(i, 2, v).font = ARIAL

ws['A9'] = 'TOTAIS GERAIS'
ws['A9'].font = ARIAL_BOLD
hdrs = ['Metrica', 'Excel (planilha)', 'Sistema (dashboard)', 'Delta Absoluto', 'Delta %']
for i, h in enumerate(hdrs):
    style_hdr(ws.cell(10, i + 1, h))

dados = [
    ('Clientes unicos', 786, 755, 755 - 786, (755 - 786) / 786),
    ('Faturamento (R$)', 9203973.81, 8523854.75, 8523854.75 - 9203973.81, (8523854.75 - 9203973.81) / 9203973.81),
]
for ri, d in enumerate(dados, start=11):
    for ci, v in enumerate(d, start=1):
        c = ws.cell(ri, ci, v)
        style_cell(c)
        if ci in (2, 3, 4):
            c.number_format = '#,##0.00' if ri == 12 else '#,##0'
        if ci == 5:
            c.number_format = '0.0%'
            c.fill = FILL_DANGER

ws['A14'] = 'DECOMPOSICAO DA DIFERENCA (-R$ 680.119,06)'
ws['A14'].font = ARIAL_BOLD
hdrs2 = ['Causa raiz', 'Delta Clientes', 'Delta Faturamento (R$)', '% do gap']
for i, h in enumerate(hdrs2):
    style_hdr(ws.cell(15, i + 1, h))

decomp = [
    ('1. Clientes ausentes da carteira RFV', -31, -241260.81, 0.355),
    ('2. Naturezas excluidas pelo filtro financial_flag=F', 0, -438858.25, 0.645),
    ('TOTAL', -31, -680119.06, 1.0),
]
for ri, d in enumerate(decomp, start=16):
    for ci, v in enumerate(d, start=1):
        c = ws.cell(ri, ci, v)
        style_cell(c)
        if ci in (2, 3):
            c.number_format = '#,##0.00'
        if ci == 4:
            c.number_format = '0.0%'
        if ri == 18:
            c.font = ARIAL_BOLD
            c.fill = FILL_LIGHT

ws['A20'] = 'COMPARATIVO POR SEGMENTO RFV'
ws['A20'].font = ARIAL_BOLD
for i, h in enumerate(df_seg.columns):
    style_hdr(ws.cell(21, i + 1, h))
for ri, row in enumerate(df_seg.itertuples(index=False), start=22):
    for ci, v in enumerate(row, start=1):
        c = ws.cell(ri, ci, v)
        style_cell(c)
        if ci > 1:
            c.alignment = RIGHT
        if ci == 4:
            if v > 0:
                c.fill = FILL_OK
            elif v < 0:
                c.fill = FILL_DANGER

ws['A34'] = 'PAUTA PROPOSTA PARA REUNIAO COM ALVES/DIEGO'
ws['A34'].font = ARIAL_BOLD
pauta = [
    '1. Filtro de natureza: definir lista canonica de nature_codes que contam como venda RFV',
    '   - Hoje: filtro dim_operation_nature.financial_flag = "F" exclui R$ 1,33 MM no periodo',
    '   - Suspeito de BUG: codigos 5101 49 e 5101 11 tem nome VENDA DE MERCADORIA mas flag != F',
    '   - Decidir: 5949 ML (REMESSA P/ DEPOSITO TEMPORARIO, R$ 495k) e venda valida?',
    '',
    '2. Carteira de clientes: hoje populada MANUALMENTE via planilhas + fuzzy match',
    '   - 31 clientes do Excel nao estao na carteira (R$ 241k)',
    '   - 4 deles foram rejeitados por fuzzy match errado',
    '   - 27 nunca foram processados (clientes novos?)',
    '',
    '3. PROPOSTA DE AUTOMACAO da carteira:',
    '   - Derivar rfv_familia direto do ERP (segmento do cliente, classe de produto, etc.)',
    '   - Ou via tag/label do CRM Pipedrive',
    '   - Eliminar dependencia de planilhas manuais',
]
for i, line in enumerate(pauta, start=35):
    c = ws.cell(i, 1, line)
    c.font = ARIAL
    if line.startswith(('1.', '2.', '3.')):
        c.font = ARIAL_BOLD

ws.column_dimensions['A'].width = 55
for col_letter in ['B', 'C', 'D', 'E']:
    ws.column_dimensions[col_letter].width = 22

# Aba 2 — Clientes Faltantes
ws2 = wb.create_sheet('Clientes Faltantes')
ws2.merge_cells('A1:F1')
ws2['A1'] = '31 CLIENTES PRESENTES NO EXCEL MAS AUSENTES DO SISTEMA - R$ 241.260,81'
ws2['A1'].font = ARIAL_TITLE
ws2['A1'].fill = FILL_PRIMARY
ws2['A1'].alignment = CENTER
ws2.row_dimensions[1].height = 30

hdrs = ['Cliente (nome do Excel)', 'Pedidos', 'Valor (R$)', 'Ultima compra', 'Status na carteira', 'Match fuzzy (se rejeitado)']
for i, h in enumerate(hdrs):
    style_hdr(ws2.cell(3, i + 1, h))

for ri, row in enumerate(faltantes.itertuples(index=False), start=4):
    cliente, pedidos, valor, ultima, bq_nome, score, match_type, status = row
    ws2.cell(ri, 1, cliente)
    ws2.cell(ri, 2, pedidos)
    ws2.cell(ri, 3, valor)
    ws2.cell(ri, 3).number_format = '#,##0.00'
    ws2.cell(ri, 4, ultima)
    ws2.cell(ri, 4).number_format = 'dd/mm/yyyy'
    ws2.cell(ri, 5, status)
    matchtxt = f'{bq_nome} (score {score:.1f}, {match_type})' if pd.notna(bq_nome) else ''
    ws2.cell(ri, 6, matchtxt)
    for ci in range(1, 7):
        ws2.cell(ri, ci).font = ARIAL
        ws2.cell(ri, ci).border = BORDER
        if ci in (2, 3, 4):
            ws2.cell(ri, ci).alignment = RIGHT
    if 'REJEITADO' in status:
        for ci in range(1, 7):
            ws2.cell(ri, ci).fill = FILL_WARN

ws2.column_dimensions['A'].width = 60
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 40
ws2.column_dimensions['F'].width = 50

# Aba 3 — Naturezas
ws3 = wb.create_sheet('Naturezas a Decidir')
ws3.merge_cells('A1:G1')
ws3['A1'] = 'NATUREZAS EXCLUIDAS PELO FILTRO financial_flag=F - R$ 1.333.641,11 NO PERIODO'
ws3['A1'].font = ARIAL_TITLE
ws3['A1'].fill = FILL_PRIMARY
ws3['A1'].alignment = CENTER
ws3.row_dimensions[1].height = 30

ws3.merge_cells('A2:G2')
ws3['A2'] = 'ACAO: marcar na coluna DECISAO ALVES/DIEGO cada natureza como INCLUIR / EXCLUIR / DUVIDOSO'
ws3['A2'].font = ARIAL_BOLD
ws3['A2'].fill = FILL_WARN
ws3['A2'].alignment = CENTER

hdrs = ['nature_code', 'Nome da natureza', 'Flag atual', 'Clientes', 'Pedidos', 'Faturamento (R$)', 'DECISAO ALVES/DIEGO']
for i, h in enumerate(hdrs):
    style_hdr(ws3.cell(4, i + 1, h))

for ri, row in enumerate(df_nat.itertuples(index=False), start=5):
    nc, nm, flag, cli, ped, fat, susp = row
    ws3.cell(ri, 1, nc)
    ws3.cell(ri, 2, nm)
    ws3.cell(ri, 3, flag)
    ws3.cell(ri, 4, int(cli))
    ws3.cell(ri, 5, int(ped))
    ws3.cell(ri, 6, fat)
    ws3.cell(ri, 6).number_format = '#,##0.00'
    ws3.cell(ri, 7, '')
    for ci in range(1, 8):
        ws3.cell(ri, ci).font = ARIAL
        ws3.cell(ri, ci).border = BORDER
        if ci in (4, 5, 6):
            ws3.cell(ri, ci).alignment = RIGHT
    if susp == 'SIM':
        for ci in range(1, 8):
            ws3.cell(ri, ci).fill = FILL_DANGER
        ws3.cell(ri, 7, 'SUSPEITO de bug no flag - nome diz VENDA')

ws3.column_dimensions['A'].width = 13
ws3.column_dimensions['B'].width = 52
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 18
ws3.column_dimensions['G'].width = 45

# Aba 4 — Drill
ws4 = wb.create_sheet('Drill Cliente x Natureza')
ws4.merge_cells('A1:F1')
ws4['A1'] = 'DRILL-DOWN: PEDIDOS DOS 755 CLIENTES EM COMUM EM NATUREZAS EXCLUIDAS (top 200) - R$ 438.858,25'
ws4['A1'].font = ARIAL_TITLE
ws4['A1'].fill = FILL_PRIMARY
ws4['A1'].alignment = CENTER
ws4.row_dimensions[1].height = 30

hdrs = ['Cliente', 'nature_code', 'Nome da natureza', 'Flag', 'Pedidos', 'Valor extra (R$)']
for i, h in enumerate(hdrs):
    style_hdr(ws4.cell(3, i + 1, h))

for ri, row in enumerate(df_extra.itertuples(index=False), start=4):
    cli, nc, nm, flag, ped, val = row
    ws4.cell(ri, 1, cli)
    ws4.cell(ri, 2, nc)
    ws4.cell(ri, 3, nm)
    ws4.cell(ri, 4, flag)
    ws4.cell(ri, 5, int(ped))
    ws4.cell(ri, 6, val)
    ws4.cell(ri, 6).number_format = '#,##0.00'
    for ci in range(1, 7):
        ws4.cell(ri, ci).font = ARIAL
        ws4.cell(ri, ci).border = BORDER
        if ci in (5, 6):
            ws4.cell(ri, ci).alignment = RIGHT

ws4.column_dimensions['A'].width = 50
ws4.column_dimensions['B'].width = 13
ws4.column_dimensions['C'].width = 50
ws4.column_dimensions['D'].width = 8
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 18

# Aba 5 — Proximos passos
ws5 = wb.create_sheet('Proximos Passos')
ws5.merge_cells('A1:C1')
ws5['A1'] = 'PROXIMOS PASSOS - A DEFINIR NA REUNIAO E EXECUTAR DEPOIS'
ws5['A1'].font = ARIAL_TITLE
ws5['A1'].fill = FILL_PRIMARY
ws5['A1'].alignment = CENTER
ws5.row_dimensions[1].height = 30

hdrs = ['#', 'Item', 'Responsavel / Notas']
for i, h in enumerate(hdrs):
    style_hdr(ws5.cell(3, i + 1, h))

passos = [
    (1, 'Validar lista de nature_codes que contam como venda RFV (preencher DECISAO da aba 3)', 'Alves + Diego'),
    (2, 'Investigar bug em dim_operation_nature: por que 5101 49 e 5101 11 tem flag != F sendo VENDA?', 'TI ERP'),
    (3, 'Decidir destino dos 31 clientes faltantes: incluir manualmente OU esperar nova rodada', 'Alves'),
    (4, 'Refazer fuzzy match com threshold mais alto OU validacao manual dos rejeitados', 'Gustavo'),
    (5, 'PROPOR AUTOMACAO da carteira RFV: derivar do ERP/CRM em vez de planilhas manuais', 'Discutir com Diego'),
    (6, 'Apos decisoes: ajustar build_silver_comercial.sql e rerodar silver_com_rfv_score', 'Gustavo'),
    (7, 'Validar nova rodada contra esta planilha e atualizar dashboard', 'Gustavo'),
    (8, 'Replicar mesmo diagnostico para familias FARMACIAS e SAC', 'Gustavo'),
]
for ri, (n, it, resp) in enumerate(passos, start=4):
    ws5.cell(ri, 1, n)
    ws5.cell(ri, 2, it)
    ws5.cell(ri, 3, resp)
    for ci in range(1, 4):
        ws5.cell(ri, ci).font = ARIAL
        ws5.cell(ri, ci).border = BORDER
        ws5.cell(ri, ci).alignment = Alignment(vertical='top', wrap_text=True,
                                                horizontal='left' if ci > 1 else 'center')
    ws5.row_dimensions[ri].height = 35

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 80
ws5.column_dimensions['C'].width = 30

out = r'C:\Users\gusta\Downloads\Auditoria_RFV_Hospitalar_Abril2026.xlsx'
wb.save(out)
print(f'OK: {out}')
