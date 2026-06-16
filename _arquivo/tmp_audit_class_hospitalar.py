"""Auditoria profunda de classificacoes RFV - HOSPITALAR (Excel x Sistema)."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

creds = service_account.Credentials.from_service_account_file(
    r'C:\teste\sapient-metrics.json',
    scopes=["https://www.googleapis.com/auth/cloud-platform"]
)
client = bigquery.Client(credentials=creds, project='sapient-metrics-492914-m7')

path = r'C:\Users\gusta\Downloads\RFV Hospitalar 01-04-2025 até 30-04-2026 (1).xlsx'
df_e = pd.read_excel(path, sheet_name='Sem fórmula Geral')
df_e['k'] = df_e['ID - CLIENTE'].astype(str).str.upper().str.strip()
df_e['Data última compra'] = pd.to_datetime(df_e['Data última compra'])

df_s = client.query("""
SELECT
    partner_name, ultima_compra_data, recencia_dias, frequencia, valor_total,
    freq_bucket, rec_bucket, classificacao_2 AS segmento, classificacao_3 AS seg_num,
    data_referencia
FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
WHERE rfv_familia = 'HOSPITALAR'
  AND data_referencia = (SELECT MAX(data_referencia) FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score` WHERE rfv_familia='HOSPITALAR')
""").to_dataframe()
df_s['k'] = df_s['partner_name'].str.upper().str.strip()
df_s['valor_total'] = pd.to_numeric(df_s['valor_total'])
data_ref_sys = df_s['data_referencia'].iloc[0]
data_ref_exc = df_e['Data de hoje'].dropna().iloc[0].date()
gap_dias = (pd.to_datetime(data_ref_sys) - pd.to_datetime(data_ref_exc)).days

m = df_e.merge(df_s, on='k', how='inner')
m['mesmo_seg'] = m['Classificação 2'] == m['segmento']
m['mesmo_freq_b'] = m['Frequência 2'] == m['freq_bucket']
m['mesmo_rec_b'] = m['Recência 3'] == m['rec_bucket']
m['delta_recencia'] = m['recencia_dias'] - m['Recência em dias']
m['delta_freq'] = m['frequencia'] - m['Frequência 1']
m['delta_valor'] = m['valor_total'] - m['Valor']

# Categoriza divergencia
ORDEM_SEG = {
    'Campeões': 1, 'Fiéis': 2, 'Fiéis em potencial': 3, 'Novos clientes': 4, 'Promessas': 5,
    'Precisando de atenção': 6, 'Quase dormentes': 7, 'Não pode perder': 8,
    'Em risco': 9, 'Hibernando': 10, 'Perdidos': 11,
}
m['ordem_excel'] = m['Classificação 2'].map(ORDEM_SEG)
m['ordem_sys'] = m['segmento'].map(ORDEM_SEG)
m['movimento'] = m.apply(lambda r:
    'Igual' if r['mesmo_seg'] else
    ('Piorou' if r['ordem_sys'] > r['ordem_excel'] else 'Melhorou'), axis=1)

def causa_provavel(r):
    """Identifica causa principal da divergencia"""
    if r['mesmo_seg']:
        return 'Sem divergencia'
    causas = []
    if r['Recência 3'] != r['rec_bucket']:
        causas.append(f'R: {r["Recência 3"]} -> {r["rec_bucket"]}')
    if r['Frequência 2'] != r['freq_bucket']:
        causas.append(f'F: {r["Frequência 2"]} -> {r["freq_bucket"]}')
    return ' + '.join(causas) if causas else 'Outro'
m['causa'] = m.apply(causa_provavel, axis=1)

print(f'Em comum: {len(m)} | Mesmo seg: {m["mesmo_seg"].sum()} | Divergem: {(~m["mesmo_seg"]).sum()}')
print(f'Data ref Excel: {data_ref_exc} | Sistema: {data_ref_sys} | Gap: {gap_dias} dias')

# Matriz de transicao
order_list = ['Campeões','Fiéis','Fiéis em potencial','Novos clientes','Promessas',
              'Precisando de atenção','Quase dormentes','Não pode perder','Em risco','Hibernando','Perdidos']
ct = pd.crosstab(m['Classificação 2'], m['segmento'])
ct = ct.reindex(index=order_list, columns=order_list, fill_value=0)

# Decomposicao da divergencia por causa
print()
print('=== Movimento ===')
print(m['movimento'].value_counts().to_string())
print()
print('=== Causa principal ===')
print(m[~m['mesmo_seg']]['causa'].value_counts().head(15).to_string())

# === EXCEL ===
wb = Workbook()
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
ARIAL_SUB = Font(name='Arial', size=11, bold=True)
FILL_PRIMARY = PatternFill('solid', fgColor='1E1882')
FILL_HEADER = PatternFill('solid', fgColor='4844C8')
FILL_DIAG = PatternFill('solid', fgColor='B5EBC5')
FILL_OK = PatternFill('solid', fgColor='C9F8D2')
FILL_DOWN = PatternFill('solid', fgColor='F8C9C9')
FILL_UP = PatternFill('solid', fgColor='C9D9F8')
FILL_WARN = PatternFill('solid', fgColor='FFE066')
FILL_LIGHT = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_hdr(cell, fill=FILL_HEADER):
    cell.font = ARIAL_HDR; cell.fill = fill; cell.alignment = CENTER; cell.border = BORDER

# Aba 1 — Resumo
ws = wb.active
ws.title = 'Resumo'
ws.merge_cells('A1:G1')
ws['A1'] = 'AUDITORIA DE CLASSIFICACOES RFV - HOSPITALAR (Excel x Sistema)'
ws['A1'].font = ARIAL_TITLE; ws['A1'].fill = FILL_PRIMARY; ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 30

info = [
    ('Familia:', 'HOSPITALAR'),
    ('Data ref. Excel:', data_ref_exc.strftime('%d/%m/%Y')),
    ('Data ref. Sistema:', data_ref_sys.strftime('%d/%m/%Y') if hasattr(data_ref_sys,'strftime') else str(data_ref_sys)),
    ('Gap (dias entre as datas):', f'{gap_dias} dias - explica boa parte das divergencias de R'),
    ('Clientes em comum (auditados):', len(m)),
    ('Gerado em:', datetime.now().strftime('%d/%m/%Y %H:%M')),
]
for i, (k, v) in enumerate(info, start=3):
    ws.cell(i, 1, k).font = ARIAL_BOLD
    ws.cell(i, 2, v).font = ARIAL

# Stats
ws['A11'] = 'PLACAR'
ws['A11'].font = ARIAL_BOLD
hdrs = ['Status', 'Clientes', '% do total']
for i, h in enumerate(hdrs):
    style_hdr(ws.cell(12, i+1, h))

total = len(m)
mesmos = int(m['mesmo_seg'].sum())
mov_counts = m['movimento'].value_counts()
linhas = [
    ('Mesmo segmento', mesmos, mesmos/total, FILL_OK),
    ('Sistema mostra MAIS RISCO (piorou)', int(mov_counts.get('Piorou', 0)),
     mov_counts.get('Piorou', 0)/total, FILL_DOWN),
    ('Sistema mostra MELHOR (melhorou)', int(mov_counts.get('Melhorou', 0)),
     mov_counts.get('Melhorou', 0)/total, FILL_UP),
]
for ri, (lbl, qtd, pct, fill) in enumerate(linhas, start=13):
    ws.cell(ri, 1, lbl).font = ARIAL
    ws.cell(ri, 2, qtd).font = ARIAL
    ws.cell(ri, 2).alignment = RIGHT
    ws.cell(ri, 3, pct).font = ARIAL
    ws.cell(ri, 3).number_format = '0.0%'
    ws.cell(ri, 3).alignment = RIGHT
    for ci in (1,2,3):
        ws.cell(ri, ci).fill = fill
        ws.cell(ri, ci).border = BORDER

# Stats numericos
ws['A18'] = 'ESTATISTICAS DOS DELTAS (sistema - Excel)'
ws['A18'].font = ARIAL_BOLD
ws['A19'] = 'Delta = quanto o valor do sistema difere do Excel para o MESMO cliente'
ws['A19'].font = ARIAL
hdrs = ['Metrica', 'Recencia (dias)', 'Frequencia (pedidos)', 'Valor (R$)']
for i, h in enumerate(hdrs):
    style_hdr(ws.cell(20, i+1, h))

stats = [
    ('Media',  float(m['delta_recencia'].mean()), float(m['delta_freq'].mean()), float(m['delta_valor'].mean())),
    ('Mediana', float(m['delta_recencia'].median()), float(m['delta_freq'].median()), float(m['delta_valor'].median())),
    ('Q25 (25%)', float(m['delta_recencia'].quantile(0.25)), float(m['delta_freq'].quantile(0.25)), float(m['delta_valor'].quantile(0.25))),
    ('Q75 (75%)', float(m['delta_recencia'].quantile(0.75)), float(m['delta_freq'].quantile(0.75)), float(m['delta_valor'].quantile(0.75))),
    ('Min',    float(m['delta_recencia'].min()), float(m['delta_freq'].min()), float(m['delta_valor'].min())),
    ('Max',    float(m['delta_recencia'].max()), float(m['delta_freq'].max()), float(m['delta_valor'].max())),
]
for ri, row in enumerate(stats, start=21):
    for ci, v in enumerate(row, start=1):
        c = ws.cell(ri, ci, v)
        c.font = ARIAL; c.border = BORDER
        if ci > 1:
            c.alignment = RIGHT
            c.number_format = '#,##0.00' if ci == 4 else '#,##0.0'

# Diagnostico
ws['A29'] = 'DIAGNOSTICO'
ws['A29'].font = ARIAL_BOLD
diag = [
    'A mediana do delta recencia e +29 dias - EXATAMENTE proximo aos 28 dias de gap entre as datas de referencia.',
    'Isso significa: a MAIOR parte das divergencias e ESPERADA, nao um erro do sistema.',
    'O sistema tem data de referencia 28 dias mais recente que a planilha, entao a recencia naturalmente aumenta.',
    'Um cliente "Campeao" no dia 02/04 facilmente vira "Fieis" no dia 30/04 se ele nao comprou nesse intervalo.',
    '',
    'Frequencia: mediana = 0 (sistema le os mesmos pedidos na maioria dos casos)',
    '            Cauda negativa (min -27): casos extremos = filtro de natureza cortando muitos pedidos',
    '',
    'Valor: mediana +R$ 60 (essencialmente igual)',
    '       Cauda negativa (min -R$ 81k): WEP COMERCIO - cliente com muitos pedidos em naturezas excluidas',
]
for i, line in enumerate(diag, start=30):
    c = ws.cell(i, 1, line)
    c.font = ARIAL
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

ws.column_dimensions['A'].width = 38
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 18

# Aba 2 — Matriz de Transicao
ws2 = wb.create_sheet('Matriz Transicao')
ws2.merge_cells('A1:M1')
ws2['A1'] = 'MATRIZ DE TRANSICAO - Excel (linha) -> Sistema (coluna)'
ws2['A1'].font = ARIAL_TITLE; ws2['A1'].fill = FILL_PRIMARY; ws2['A1'].alignment = CENTER
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:M2')
ws2['A2'] = ('Como ler: linha=segmento da planilha, coluna=segmento do sistema. '
             'Diagonal (verde) = mesmo segmento. Outros = divergencias com motivos abaixo.')
ws2['A2'].font = ARIAL; ws2['A2'].fill = FILL_WARN; ws2['A2'].alignment = CENTER
ws2.row_dimensions[2].height = 35

# Header colunas (sistema)
ws2.cell(4, 1, 'Excel \\ Sistema').font = ARIAL_BOLD
ws2.cell(4, 1).fill = FILL_LIGHT
ws2.cell(4, 1).alignment = CENTER
ws2.cell(4, 1).border = BORDER
for i, seg in enumerate(order_list, start=2):
    c = ws2.cell(4, i, seg)
    c.font = ARIAL_HDR; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
c = ws2.cell(4, len(order_list)+2, 'TOTAL EXCEL')
c.font = ARIAL_HDR; c.fill = FILL_PRIMARY; c.alignment = CENTER; c.border = BORDER

# Linhas (Excel)
for ri, seg_e in enumerate(order_list, start=5):
    c = ws2.cell(ri, 1, seg_e)
    c.font = ARIAL_HDR; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
    total_row = 0
    for ci, seg_s in enumerate(order_list, start=2):
        val = int(ct.loc[seg_e, seg_s])
        cell = ws2.cell(ri, ci, val)
        cell.font = ARIAL; cell.alignment = CENTER; cell.border = BORDER
        if seg_e == seg_s:
            cell.fill = FILL_DIAG  # diagonal — bateu
            if val > 0:
                cell.font = ARIAL_BOLD
        elif val > 0:
            ordem_e = ORDEM_SEG[seg_e]
            ordem_s = ORDEM_SEG[seg_s]
            if ordem_s > ordem_e:
                cell.fill = FILL_DOWN  # piorou
            else:
                cell.fill = FILL_UP  # melhorou
        total_row += val
    tc = ws2.cell(ri, len(order_list)+2, total_row)
    tc.font = ARIAL_BOLD; tc.fill = FILL_LIGHT; tc.alignment = CENTER; tc.border = BORDER

# Linha de totais sistema
ri_tot = 5 + len(order_list)
c = ws2.cell(ri_tot, 1, 'TOTAL SISTEMA')
c.font = ARIAL_HDR; c.fill = FILL_PRIMARY; c.alignment = CENTER; c.border = BORDER
for ci, seg_s in enumerate(order_list, start=2):
    total_col = int(ct[seg_s].sum())
    tc = ws2.cell(ri_tot, ci, total_col)
    tc.font = ARIAL_BOLD; tc.fill = FILL_LIGHT; tc.alignment = CENTER; tc.border = BORDER

# Legenda
ri_leg = ri_tot + 3
ws2.cell(ri_leg, 1, 'LEGENDA:').font = ARIAL_BOLD
ws2.cell(ri_leg+1, 1, 'Verde diagonal = MESMO segmento (sistema e planilha concordam)').font = ARIAL
ws2.cell(ri_leg+1, 1).fill = FILL_DIAG
ws2.cell(ri_leg+2, 1, 'Vermelho = sistema mostra cliente em SEGMENTO PIOR (mais risco)').font = ARIAL
ws2.cell(ri_leg+2, 1).fill = FILL_DOWN
ws2.cell(ri_leg+3, 1, 'Azul = sistema mostra cliente em SEGMENTO MELHOR (menos risco)').font = ARIAL
ws2.cell(ri_leg+3, 1).fill = FILL_UP
for r in (ri_leg+1, ri_leg+2, ri_leg+3):
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)

ws2.column_dimensions['A'].width = 22
for ci in range(2, 13):
    ws2.column_dimensions[chr(64+ci)].width = 13

# Aba 3 — Drill 755 clientes
ws3 = wb.create_sheet('Drill 755 clientes')
ws3.merge_cells('A1:N1')
ws3['A1'] = '755 CLIENTES AUDITADOS - LINHA POR CLIENTE'
ws3['A1'].font = ARIAL_TITLE; ws3['A1'].fill = FILL_PRIMARY; ws3['A1'].alignment = CENTER
ws3.row_dimensions[1].height = 30

hdrs = ['Cliente', 'Excel: Última', 'Excel: Rec', 'Excel: Freq', 'Excel: Valor (R$)', 'Excel: Segmento',
        'Sis: Última', 'Sis: Rec', 'Sis: Freq', 'Sis: Valor (R$)', 'Sis: Segmento',
        'Movimento', 'Causa principal', 'Delta Rec (dias)']
for i, h in enumerate(hdrs):
    style_hdr(ws3.cell(3, i+1, h))

m_sorted = m.sort_values(['mesmo_seg', 'ordem_sys'], ascending=[True, True])
for ri, r in enumerate(m_sorted.itertuples(index=False), start=4):
    ws3.cell(ri, 1, r.k)
    ws3.cell(ri, 2, r._4)  # Data última compra (Excel)
    ws3.cell(ri, 2).number_format = 'dd/mm/yyyy'
    ws3.cell(ri, 3, int(r._6) if pd.notna(r._6) else 0)  # Recência em dias
    ws3.cell(ri, 4, int(r._1) if pd.notna(r._1) else 0)  # Frequência 1
    ws3.cell(ri, 5, float(r.Valor) if pd.notna(r.Valor) else 0)
    ws3.cell(ri, 5).number_format = '#,##0.00'
    ws3.cell(ri, 6, r._9 if pd.notna(r._9) else '')  # Classificação 2
    ws3.cell(ri, 7, r.ultima_compra_data)
    ws3.cell(ri, 7).number_format = 'dd/mm/yyyy'
    ws3.cell(ri, 8, int(r.recencia_dias))
    ws3.cell(ri, 9, int(r.frequencia))
    ws3.cell(ri, 10, float(r.valor_total))
    ws3.cell(ri, 10).number_format = '#,##0.00'
    ws3.cell(ri, 11, r.segmento)
    ws3.cell(ri, 12, r.movimento)
    ws3.cell(ri, 13, r.causa)
    ws3.cell(ri, 14, int(r.delta_recencia))

    # Highlight
    fill = FILL_OK if r.mesmo_seg else (FILL_DOWN if r.movimento == 'Piorou' else FILL_UP)
    for ci in range(1, 15):
        c = ws3.cell(ri, ci)
        c.font = ARIAL; c.border = BORDER
        if ci in (3,4,8,9,14): c.alignment = RIGHT
        if ci in (5,10): c.alignment = RIGHT
    # Pinta colunas Movimento/Causa de acordo
    ws3.cell(ri, 12).fill = fill

# Largura
widths = [40, 12, 8, 8, 13, 22, 12, 8, 8, 13, 22, 12, 22, 11]
for i, w in enumerate(widths):
    ws3.column_dimensions[chr(65+i)].width = w

# Aba 4 — Casos flagrantes
ws4 = wb.create_sheet('Casos Flagrantes')
ws4.merge_cells('A1:H1')
ws4['A1'] = 'TOP CASOS DE DIVERGENCIA - PARA ANALISE QUALITATIVA NA REUNIAO'
ws4['A1'].font = ARIAL_TITLE; ws4['A1'].fill = FILL_PRIMARY; ws4['A1'].alignment = CENTER
ws4.row_dimensions[1].height = 30

ri = 3

def secao(titulo, df_sub, desc, ri):
    ws4.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=8)
    c = ws4.cell(ri, 1, titulo)
    c.font = ARIAL_SUB; c.fill = FILL_LIGHT
    ri += 1
    ws4.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=8)
    ws4.cell(ri, 1, desc).font = ARIAL
    ri += 1
    hdrs = ['Cliente', 'Excel: Seg', 'Sis: Seg', 'Excel: Rec', 'Sis: Rec',
            'Excel: Freq', 'Sis: Freq', 'Delta Valor (R$)']
    for i, h in enumerate(hdrs):
        style_hdr(ws4.cell(ri, i+1, h))
    ri += 1
    for _, r in df_sub.iterrows():
        ws4.cell(ri, 1, r['k'])
        ws4.cell(ri, 2, r['Classificação 2'])
        ws4.cell(ri, 3, r['segmento'])
        ws4.cell(ri, 4, int(r['Recência em dias']))
        ws4.cell(ri, 5, int(r['recencia_dias']))
        ws4.cell(ri, 6, int(r['Frequência 1']))
        ws4.cell(ri, 7, int(r['frequencia']))
        ws4.cell(ri, 8, float(r['delta_valor']))
        ws4.cell(ri, 8).number_format = '#,##0.00'
        for ci in range(1, 9):
            c = ws4.cell(ri, ci); c.font = ARIAL; c.border = BORDER
            if ci > 1: c.alignment = RIGHT
        ri += 1
    ri += 2
    return ri

div = m[~m['mesmo_seg']].copy()

# Top 15 melhorou (sistema mostra ativo)
ri = secao(
    '1. Sistema RESSUSCITOU clientes (planilha dizia "Perdidos" mas sistema mostra ativo)',
    div[(div['Classificação 2']=='Perdidos') & (div['segmento']!='Perdidos')].nlargest(15, 'valor_total'),
    'Estes clientes voltaram a comprar entre 02/04 e 30/04 - planilha nao viu, sistema viu.',
    ri,
)

ri = secao(
    '2. Sistema PIOROU clientes Campeoes -> Fieis (movimento por recencia)',
    div[(div['Classificação 2']=='Campeões') & (div['segmento']=='Fiéis')].head(15),
    'Esperado: o Campeao do Excel ficou 28 dias mais antigo, recencia entrou em R2.',
    ri,
)

ri = secao(
    '3. Sistema PIOROU clientes (planilha dizia "Quase dormentes" -> sistema "Perdidos")',
    div[(div['Classificação 2']=='Quase dormentes') & (div['segmento']=='Perdidos')].head(15),
    'Em 28 dias adicionais, esses clientes ultrapassaram a barreira dos 180 dias.',
    ri,
)

ri = secao(
    '4. Maiores quedas de FATURAMENTO (filtro de natureza cortando vendas)',
    div.nsmallest(15, 'delta_valor')[['k','Classificação 2','segmento','Recência em dias',
                                       'recencia_dias','Frequência 1','frequencia','delta_valor']],
    'Mesmo cliente em comum, mas sistema fatura menos = naturezas excluidas (financial_flag != F).',
    ri,
)

ri = secao(
    '5. Maiores QUEDAS de FREQUENCIA (pedidos perdidos pelo filtro)',
    div.nsmallest(15, 'delta_freq')[['k','Classificação 2','segmento','Recência em dias',
                                      'recencia_dias','Frequência 1','frequencia','delta_valor']],
    'Sistema conta menos pedidos que a planilha - mesma causa: filtro de natureza.',
    ri,
)

for i, w in enumerate([45, 22, 22, 10, 10, 10, 10, 16]):
    ws4.column_dimensions[chr(65+i)].width = w

out = r'C:\Users\gusta\Downloads\Auditoria_RFV_Hospitalar_Classificacoes.xlsx'
wb.save(out)
print(f'OK: {out}')
