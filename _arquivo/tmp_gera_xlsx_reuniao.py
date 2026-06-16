import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\gusta\Downloads\RFV_BQ_vs_Alves_abril2026.xlsx"

HOSPITALAR = [
    ("Campeões",           45,  80, 3851737, 5433495),
    ("Fiéis",              61,  47, 1234024,  913016),
    ("Fiéis em potencial", 58, 112,  475859,  737901),
    ("Novos clientes",      9,  38,   13833,   90329),
    ("Quase dormentes",    61,  63,  185711,  279801),
    ("Não pode perder",     8,  11,  160898,  332501),
    ("Em risco",           31,  21,  245746,  152625),
    ("Hibernando",         23,  41,  191514,  256045),
    ("Perdidos",          303, 339,  924009,  908599),
]
HOSPITALAR_TOTAL = ("TOTAL", 629, 786, 7382300, 9203973)

FARMACIAS = [
    ("Campeões–Fiéis–FP",  45,   0, 132273,      0),
    ("Em risco",           28,  31,  75307,  96340),
    ("Hibernando",         10,  21,  26594,  23562),
    ("Perdidos",          147, 191, 174738, 231345),
]
FARMACIAS_TOTAL = ("TOTAL", 248, 248, 446183, 375143)

SAC = []
SAC_TOTAL = ("TOTAL", 76, 79, 197009, 221180)

PONTOS = [
    ("1. RFV GERAL — falta canal",
     "Soma dos 3 canais consultados (Hospitalar + Farmácias + SAC) dá 953 (BQ) vs 1.113 (Alves). "
     "A RFV geral deveria passar de 2.000 clientes — provavelmente o canal MARKETPLACE não entrou "
     "na consulta. Validar antes da reunião e/ou alinhar na pauta."),
    ("2. HOSPITALAR",
     "−157 clientes vs Alves — já explicado: redistribuição de carteira (Kauan Ramos + SAC). "
     "Não é perda de dado, é reorganização."),
    ("3. FARMÁCIAS",
     "Mesmo total (248), faturamento +R$ 71K a mais no BQ — nossos thresholds são mais generosos "
     "que os do Alves (ele usava R1≤90 dias, nós usamos R1≤30 dias). Por isso temos Campeões/Fiéis "
     "e ele não tinha. Alinhar thresholds na reunião."),
    ("4. SAC",
     "Praticamente igual — diferença de 3 clientes e −R$ 24K, irrelevante."),
]

GERAL = [
    ("Hospitalar",  629,  786, 7382300, 9203973),
    ("Farmácias",   248,  248,  446183,  375143),
    ("SAC",          76,   79,  197009,  221180),
]
GERAL_TOTAL = ("TOTAL (3 canais)", 953, 1113, 8025492, 9800296)

# estilos
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL  = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT  = Font(bold=True)
TITLE_FONT  = Font(bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

COLS = ["Segmento", "BQ (clientes)", "Alves (clientes)", "Δ Clientes",
        "BQ Faturamento", "Alves Faturamento", "Δ Faturamento"]

def write_sheet(ws, titulo, linhas, total):
    ws["A1"] = titulo
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A2"] = "Janela: abr/2025 – abr/2026  •  Comparativo BQ vs planilha Alves"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    # header
    for j, c in enumerate(COLS, start=1):
        cell = ws.cell(row=4, column=j, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    r = 5
    for seg, bq, al, bqf, alf in linhas:
        ws.cell(r, 1, seg).alignment = LEFT
        ws.cell(r, 2, bq).alignment = CENTER
        ws.cell(r, 3, al).alignment = CENTER
        ws.cell(r, 4, bq - al).alignment = CENTER
        ws.cell(r, 5, bqf).number_format = 'R$ #,##0'
        ws.cell(r, 6, alf).number_format = 'R$ #,##0'
        delta = bqf - alf
        cell_delta = ws.cell(r, 7, delta)
        cell_delta.number_format = 'R$ #,##0;[Red]-R$ #,##0'
        for j in range(1, 8):
            ws.cell(r, j).border = BORDER
        r += 1

    # total
    seg, bq, al, bqf, alf = total
    ws.cell(r, 1, seg).alignment = LEFT
    ws.cell(r, 2, bq).alignment = CENTER
    ws.cell(r, 3, al).alignment = CENTER
    ws.cell(r, 4, bq - al).alignment = CENTER
    ws.cell(r, 5, bqf).number_format = 'R$ #,##0'
    ws.cell(r, 6, alf).number_format = 'R$ #,##0'
    cell_delta = ws.cell(r, 7, bqf - alf)
    cell_delta.number_format = 'R$ #,##0;[Red]-R$ #,##0'
    for j in range(1, 8):
        c = ws.cell(r, j)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT
        c.border = BORDER

    # widths
    widths = [22, 14, 18, 12, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def write_pontos(ws):
    ws["A1"] = "Pontos para a reunião"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A3"] = "Tópico"
    ws["B3"] = "Mensagem"
    for c in ("A3", "B3"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
        ws[c].alignment = CENTER
        ws[c].border = BORDER

    r = 4
    for topico, msg in PONTOS:
        ws.cell(r, 1, topico).font = TOTAL_FONT
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="top")
        ws.cell(r, 2, msg).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for j in (1, 2):
            ws.cell(r, j).border = BORDER
        ws.row_dimensions[r].height = 60
        r += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


wb = Workbook()
wb.remove(wb.active)

ws0 = wb.create_sheet("Geral")
ws0["A1"] = "RFV GERAL — consolidado por canal (abril/2026)"
ws0["A1"].font = TITLE_FONT
ws0.merge_cells("A1:G1")
ws0["A2"] = "Atenção: soma cobre apenas Hospitalar + Farmácias + SAC. Canal MARKETPLACE provavelmente ausente da consulta — esperado >2.000 clientes no total."
ws0["A2"].font = Font(italic=True, size=10, color="C00000")
ws0.merge_cells("A2:G2")
ws0.row_dimensions[2].height = 30
ws0["A2"].alignment = Alignment(wrap_text=True, vertical="center")

for j, c in enumerate(["Canal", "BQ (clientes)", "Alves (clientes)", "Δ Clientes",
                       "BQ Faturamento", "Alves Faturamento", "Δ Faturamento"], start=1):
    cell = ws0.cell(row=4, column=j, value=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

r = 5
for canal, bq, al, bqf, alf in GERAL:
    ws0.cell(r, 1, canal).alignment = LEFT
    ws0.cell(r, 2, bq).alignment = CENTER
    ws0.cell(r, 3, al).alignment = CENTER
    ws0.cell(r, 4, bq - al).alignment = CENTER
    ws0.cell(r, 5, bqf).number_format = 'R$ #,##0'
    ws0.cell(r, 6, alf).number_format = 'R$ #,##0'
    ws0.cell(r, 7, bqf - alf).number_format = 'R$ #,##0;[Red]-R$ #,##0'
    for j in range(1, 8):
        ws0.cell(r, j).border = BORDER
    r += 1

canal, bq, al, bqf, alf = GERAL_TOTAL
ws0.cell(r, 1, canal).alignment = LEFT
ws0.cell(r, 2, bq).alignment = CENTER
ws0.cell(r, 3, al).alignment = CENTER
ws0.cell(r, 4, bq - al).alignment = CENTER
ws0.cell(r, 5, bqf).number_format = 'R$ #,##0'
ws0.cell(r, 6, alf).number_format = 'R$ #,##0'
ws0.cell(r, 7, bqf - alf).number_format = 'R$ #,##0;[Red]-R$ #,##0'
for j in range(1, 8):
    c = ws0.cell(r, j)
    c.fill = TOTAL_FILL
    c.font = TOTAL_FONT
    c.border = BORDER

r += 2
ws0.cell(r, 1, "⚠ Marketplace não consultado").font = Font(bold=True, color="C00000")
ws0.cell(r+1, 1, "Esperado RFV geral > 2.000 clientes. Validar inclusão do canal Marketplace.").font = Font(italic=True, color="595959")
ws0.merge_cells(start_row=r,   start_column=1, end_row=r,   end_column=7)
ws0.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=7)

widths = [22, 14, 18, 12, 18, 18, 18]
for i, w in enumerate(widths, start=1):
    ws0.column_dimensions[get_column_letter(i)].width = w
ws0.freeze_panes = "A5"

ws1 = wb.create_sheet("Hospitalar")
write_sheet(ws1, "HOSPITALAR — BQ vs Alves (abril/2026)", HOSPITALAR, HOSPITALAR_TOTAL)

ws2 = wb.create_sheet("Farmácias")
write_sheet(ws2, "FARMÁCIAS — BQ vs Alves (abril/2026)", FARMACIAS, FARMACIAS_TOTAL)

ws3 = wb.create_sheet("SAC")
write_sheet(ws3, "SAC — BQ vs Alves (abril/2026)", SAC, SAC_TOTAL)

ws4 = wb.create_sheet("Pontos para reunião")
write_pontos(ws4)

wb.save(OUT)
print(f"OK → {OUT}")
