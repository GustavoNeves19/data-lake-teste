"""Auditoria RFV SAC para reuniao Alves/Diego."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

creds = service_account.Credentials.from_service_account_file(
    r'C:\teste\sapient-metrics.json',
    scopes=["https://www.googleapis.com/auth/cloud-platform"]
)
client = bigquery.Client(credentials=creds, project='sapient-metrics-492914-m7')

path = r'C:\Users\gusta\Downloads\RFV SAC 01-04-2025 até 30-04-2026 (1).xlsx'
df_e = pd.read_excel(path, sheet_name='Base Inicial').dropna(subset=['CLIENTE'])
df_e['VALOR'] = pd.to_numeric(df_e['VALOR'])
ex_g = df_e.groupby('CLIENTE').agg(
    pedidos_brutos=('DATA', 'count'),
    valor=('VALOR', 'sum'),
    ultima=('DATA', 'max'),
).reset_index()
ex_g['k'] = ex_g['CLIENTE'].str.upper().str.strip()

df_s = client.query("""
SELECT c.partner_name AS cliente, COUNT(DISTINCT o.order_number) AS pedidos,
       ROUND(SUM(o.total_amount), 2) AS valor
FROM `sapient-metrics-492914-m7.dm_orders.fact_sales_order` o
JOIN `sapient-metrics-492914-m7.silver_comercial.param_com_rfv_carteira` c
  ON c.partner_code = o.partner_code AND c.is_active = TRUE
  AND c.salesperson_name NOT IN ('Eduardo', 'Karina')
JOIN `sapient-metrics-492914-m7.dm_orders.dim_operation_nature` n
  ON n.nature_code = o.nature_code AND n.financial_flag = 'F'
WHERE o.order_status IN (3, 4) AND c.rfv_familia = 'SAC'
  AND o.order_date BETWEEN '2025-04-01' AND '2026-04-30'
GROUP BY 1
""").to_dataframe()
df_s['k'] = df_s['cliente'].str.upper().str.strip()
df_s['valor'] = pd.to_numeric(df_s['valor'])

so_excel = ex_g[~ex_g['k'].isin(df_s['k'])].copy()
m = ex_g.merge(df_s[['k', 'valor', 'pedidos']].rename(
    columns={'valor': 'valor_s', 'pedidos': 'pedidos_s'}), on='k', how='inner')
m['delta_valor'] = m['valor_s'] - m['valor']
m['delta_pedidos'] = m['pedidos_s'] - m['pedidos_brutos']

# Duplicatas
df_e_sorted = df_e.sort_values(['CLIENTE', 'DATA', 'VALOR']).reset_index(drop=True)
df_e_sorted['dup'] = df_e_sorted.duplicated(subset=['CLIENTE', 'DATA', 'VALOR'], keep=False)
df_dups = df_e_sorted[df_e_sorted['dup']].copy()

# Naturezas excluidas para SAC
df_nat = client.query("""
SELECT
  o.nature_code, ANY_VALUE(n.nature_name) AS nat_name,
  ANY_VALUE(n.financial_flag) AS financial_flag,
  COUNT(DISTINCT c.partner_name) AS clientes,
  COUNT(DISTINCT o.order_number) AS pedidos,
  ROUND(SUM(o.total_amount), 2) AS faturamento
FROM `sapient-metrics-492914-m7.dm_orders.fact_sales_order` o
JOIN `sapient-metrics-492914-m7.silver_comercial.param_com_rfv_carteira` c
  ON c.partner_code = o.partner_code AND c.is_active = TRUE
  AND c.salesperson_name NOT IN ('Eduardo', 'Karina')
LEFT JOIN `sapient-metrics-492914-m7.dm_orders.dim_operation_nature` n
  ON n.nature_code = o.nature_code
WHERE o.order_status IN (3, 4) AND c.rfv_familia = 'SAC'
  AND o.order_date BETWEEN '2025-04-01' AND '2026-04-30'
  AND (n.financial_flag != 'F' OR n.financial_flag IS NULL)
GROUP BY 1
ORDER BY faturamento DESC
""").to_dataframe()
df_nat['faturamento'] = pd.to_numeric(df_nat['faturamento'])

df_seg_sys = client.query("""
SELECT classificacao_3 AS seg_num, ANY_VALUE(classificacao_2) AS segmento,
       COUNT(DISTINCT partner_name) AS clientes,
       ROUND(SUM(valor_total), 2) AS faturamento
FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
WHERE rfv_familia = 'SAC'
  AND data_referencia = (SELECT MAX(data_referencia)
                         FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
                         WHERE rfv_familia = 'SAC')
GROUP BY 1 ORDER BY 1
""").to_dataframe()

# Comparativo segmento (Excel valores do Resumo RFV)
seg_excel = {
    1: ('Campeoes', 0, 0),
    2: ('Fieis', 0, 0),
    3: ('Fieis em potencial', 1, 630.0),
    4: ('Novos clientes', 0, 0),
    5: ('Promessas', 3, 612.54),
    6: ('Precisando de atencao', 0, 0),
    7: ('Quase dormentes', 8, 4664.13),
    8: ('Nao pode perder', 4, 102913.68),
    9: ('Em risco', 4, 26242.02),
    10: ('Hibernando', 4, 47362.32),
    11: ('Perdidos', 55, 38755.30),
}
seg_sys = {int(r['seg_num']): (str(r['segmento']), int(r['clientes']), float(r['faturamento']))
           for _, r in df_seg_sys.iterrows()}
comp = []
for seg_num, (nome, exc_cli, exc_fat) in seg_excel.items():
    _, sys_cli, sys_fat = seg_sys.get(seg_num, (nome, 0, 0.0))
    comp.append([nome, exc_cli, sys_cli, sys_cli - exc_cli, exc_fat, sys_fat, sys_fat - exc_fat])
df_comp = pd.DataFrame(comp, columns=['Segmento', 'Excel Cli', 'Sis Cli', 'Delta Cli',
                                       'Excel Fat (R$)', 'Sis Fat (R$)', 'Delta Fat (R$)'])

# ====== EXCEL ======
wb = Workbook()
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FILL_PRIMARY = PatternFill('solid', fgColor='1E1882')
FILL_HEADER = PatternFill('solid', fgColor='4844C8')
FILL_WARN = PatternFill('solid', fgColor='FFE066')
FILL_DANGER = PatternFill('solid', fgColor='F8C9C9')
FILL_OK = PatternFill('solid', fgColor='C9F8D2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_hdr(cell, fill=FILL_HEADER):
    cell.font = ARIAL_HDR
    cell.fill = fill
    cell.alignment = CENTER
    cell.border = BORDER


def style_cell(cell, fmt=None):
    cell.font = ARIAL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


# Aba 1
ws = wb.active
ws.title = 'Resumo Executivo'
ws.merge_cells('A1:G1')
ws['A1'] = 'AUDITORIA RFV SAC - Excel (planilha) x Sistema (Dashboard Nevoni)'
ws['A1'].font = ARIAL_TITLE
ws['A1'].fill = FILL_PRIMARY
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 30

info = [
    ('Periodo auditado:', '01/04/2025 a 30/04/2026'),
    ('Familia RFV:', 'SAC'),
    ('Fonte Excel:', 'RFV SAC 01-04-2025 ate 30-04-2026 (1).xlsx'),
    ('Fonte Sistema:', 'silver_comercial.silver_com_rfv_score'),
    ('Gerado em:', datetime.now().strftime('%d/%m/%Y %H:%M')),
]
for i, (k, v) in enumerate(info, start=3):
    ws.cell(i, 1, k).font = ARIAL_BOLD
    ws.cell(i, 2, v).font = ARIAL

ws.merge_cells('A9:G9')
ws['A9'] = 'ATENCAO: planilha SAC tem os mesmos problemas estruturais da planilha Farmacias (vide aba "Achados Excel")'
ws['A9'].font = ARIAL_BOLD
ws['A9'].fill = FILL_WARN
ws['A9'].alignment = CENTER

ws['A11'] = 'TOTAIS GERAIS'
ws['A11'].font = ARIAL_BOLD
hdrs = ['Metrica', 'Excel (planilha)', 'Sistema (dashboard)', 'Delta', 'Delta %']
for i, h in enumerate(hdrs):
    style_hdr(ws.cell(12, i + 1, h))

ex_total_cli = 76  # base inicial agregada; Resumo do Excel diz 79
sys_total_cli = len(df_s)
ex_total_fat = float(ex_g['valor'].sum())
sys_total_fat = float(df_s['valor'].sum())

dados = [
    ('Clientes unicos (base inicial)', ex_total_cli, sys_total_cli,
     sys_total_cli - ex_total_cli, (sys_total_cli - ex_total_cli) / ex_total_cli),
    ('Clientes (Resumo RFV do Excel)', 79, sys_total_cli, sys_total_cli - 79, (sys_total_cli - 79) / 79),
    ('Faturamento (R$)', ex_total_fat, sys_total_fat,
     sys_total_fat - ex_total_fat, (sys_total_fat - ex_total_fat) / ex_total_fat),
]
for ri, d in enumerate(dados, start=13):
    for ci, v in enumerate(d, start=1):
        c = ws.cell(ri, ci, v)
        style_cell(c)
        if ci in (2, 3, 4):
            c.number_format = '#,##0.00' if ri == 15 else '#,##0'
        if ci == 5:
            c.number_format = '0.0%'
            c.fill = FILL_OK if v > 0 else FILL_DANGER

# Comparativo segmentos
ws['A17'] = 'COMPARATIVO POR SEGMENTO RFV'
ws['A17'].font = ARIAL_BOLD
ws.merge_cells('A18:G18')
ws['A18'] = 'Nota: Excel SAC tambem tem ZERO em 4 segmentos topo (Campeoes/Fieis/Novos/Precisando). Mesmo padrao Farmacias.'
ws['A18'].font = ARIAL
ws['A18'].fill = FILL_WARN

for i, h in enumerate(df_comp.columns):
    style_hdr(ws.cell(19, i + 1, h))
for ri, row in enumerate(df_comp.itertuples(index=False), start=20):
    for ci, v in enumerate(row, start=1):
        c = ws.cell(ri, ci, v)
        style_cell(c)
        if ci > 1:
            c.alignment = RIGHT
        if ci in (5, 6, 7):
            c.number_format = '#,##0.00'
        if ci == 4:
            if v > 0:
                c.fill = FILL_OK
            elif v < 0:
                c.fill = FILL_DANGER
        if ci == 7:
            if v > 0:
                c.fill = FILL_OK
            elif v < 0:
                c.fill = FILL_DANGER

# Pauta
ws['A33'] = 'PONTOS PARA REUNIAO COM ALVES/DIEGO'
ws['A33'].font = ARIAL_BOLD
pauta = [
    '1. Planilha de SAC apresenta os MESMOS problemas estruturais da Farmacias:',
    '   - Diversos #N/A nas formulas da matriz "Resultado RFV" (linhas 4, 5, 8, 13, 14)',
    '   - Segmentos topo (Campeoes/Fieis/Novos/Precisando) com 0 clientes - sugere erro de formula',
    '   - Discrepancia entre Total Geral da matriz (79) e Base Inicial agregada (76)',
    '',
    '2. Compare numerico Excel x Sistema:',
    f'   - Volume total: Excel 76-79 / Sistema {sys_total_cli} clientes',
    f'   - Faturamento: Sistema R$ {sys_total_fat:,.2f} vs Excel R$ {ex_total_fat:,.2f}',
    '   - Sistema apresenta distribuicao por segmento que faz mais sentido de negocio',
    '   - 10 clientes do Excel nao estao no sistema (R$ 58k)',
    '',
    '3. Filtro de natureza para SAC:',
    f'   - Naturezas excluidas pelo financial_flag=F: {len(df_nat)} naturezas, R$ {df_nat["faturamento"].sum():,.2f}',
    '   - Mesma discussao do Hospitalar: validar lista canonica',
    '',
    '4. Decisoes necessarias:',
    '   - Sistema vira referencia para SAC (assumindo planilha bugada)?',
    '   - Ou reconstruir a planilha SAC antes de comparar?',
    '   - Validar 10 clientes ausentes - sao da carteira SAC ou de outras familias?',
    '',
    '5. AUTOMACAO (reforco da pauta Hospitalar/Farmacias):',
    '   - Tres planilhas, tres rodadas de problemas diferentes',
    '   - Necessidade urgente de eliminar dependencia de planilhas manuais',
    '   - Carteira RFV precisa ser derivada do ERP/CRM automaticamente',
]
for i, line in enumerate(pauta, start=34):
    c = ws.cell(i, 1, line)
    c.font = ARIAL
    if line.startswith(('1.', '2.', '3.', '4.', '5.')):
        c.font = ARIAL_BOLD

ws.column_dimensions['A'].width = 60
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col_letter].width = 16

# Aba 2 - Achados
ws2 = wb.create_sheet('Achados Excel')
ws2.merge_cells('A1:D1')
ws2['A1'] = 'PROBLEMAS IDENTIFICADOS NA PLANILHA DE SAC'
ws2['A1'].font = ARIAL_TITLE
ws2['A1'].fill = FILL_PRIMARY
ws2['A1'].alignment = CENTER
ws2.row_dimensions[1].height = 30

hdrs = ['#', 'Achado', 'Evidencia', 'Impacto']
for i, h in enumerate(hdrs):
    style_hdr(ws2.cell(3, i + 1, h))

achados = [
    ('1', 'Celulas #N/A na matriz Resultado RFV',
     'Aba "Resultado RFV" linhas 4, 5, 8, 13, 14: contagens #N/A',
     'Matriz inutilizavel - 4 segmentos topo com 0 clientes'),
    ('2', 'Discrepancia: Base Inicial (76 clientes) x Resumo RFV (79 clientes)',
     'Total Geral linha 30 da matriz: 79 clientes. Base Inicial agregada: 76',
     '3 clientes "fantasmas" no Resumo, fonte nao identificada'),
    ('3', 'Possiveis duplicatas',
     f'{len(df_dups)} linhas suspeitas (mesma chave CLIENTE+DATA+VALOR)',
     'Inflacionamento de count' if len(df_dups) > 0 else 'Nenhuma duplicata exata'),
    ('4', 'Aba "Remover duplicatas" com 1000 linhas (template/vazia)',
     'Aba presente mas nao processada',
     'Sugere que o dedup nao foi executado'),
    ('5', '55 de 76 clientes em "Perdidos" (72%)',
     'Linha 22 da matriz: 55 clientes',
     'Distribuicao improvavel - sugere problema de calculo ou regua restritiva'),
]
for ri, ach in enumerate(achados, start=4):
    for ci, v in enumerate(ach, start=1):
        c = ws2.cell(ri, ci, v)
        c.font = ARIAL
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical='top',
                                horizontal='center' if ci == 1 else 'left')
    ws2.row_dimensions[ri].height = 50

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 38
ws2.column_dimensions['C'].width = 55
ws2.column_dimensions['D'].width = 38

# Aba 3 - Clientes Faltantes
ws3 = wb.create_sheet('Clientes Faltantes')
ws3.merge_cells('A1:D1')
ws3['A1'] = f'{len(so_excel)} CLIENTES PRESENTES NO EXCEL MAS AUSENTES DO SISTEMA - R$ {so_excel["valor"].sum():,.2f}'
ws3['A1'].font = ARIAL_TITLE
ws3['A1'].fill = FILL_PRIMARY
ws3['A1'].alignment = CENTER
ws3.row_dimensions[1].height = 30

hdrs = ['Cliente (Excel)', 'Pedidos', 'Valor (R$)', 'Ultima compra']
for i, h in enumerate(hdrs):
    style_hdr(ws3.cell(3, i + 1, h))

for ri, row in enumerate(so_excel.sort_values('valor', ascending=False).itertuples(index=False), start=4):
    cliente, pedidos, valor, ultima, k = row
    ws3.cell(ri, 1, cliente)
    ws3.cell(ri, 2, int(pedidos))
    ws3.cell(ri, 3, valor)
    ws3.cell(ri, 3).number_format = '#,##0.00'
    ws3.cell(ri, 4, ultima)
    ws3.cell(ri, 4).number_format = 'dd/mm/yyyy'
    for ci in range(1, 5):
        ws3.cell(ri, ci).font = ARIAL
        ws3.cell(ri, ci).border = BORDER
        if ci > 1:
            ws3.cell(ri, ci).alignment = RIGHT

ws3.column_dimensions['A'].width = 55
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 14

# Aba 4 - Drill clientes em comum
ws4 = wb.create_sheet('Drill em comum')
ws4.merge_cells('A1:G1')
ws4['A1'] = f'COMPARATIVO PEDIDOS x VALOR - {len(m)} CLIENTES EM COMUM'
ws4['A1'].font = ARIAL_TITLE
ws4['A1'].fill = FILL_PRIMARY
ws4['A1'].alignment = CENTER
ws4.row_dimensions[1].height = 30

hdrs = ['Cliente', 'Excel Pedidos', 'Sis Pedidos', 'Excel Valor (R$)', 'Sis Valor (R$)',
        'Delta Pedidos', 'Delta Valor (R$)']
for i, h in enumerate(hdrs):
    style_hdr(ws4.cell(3, i + 1, h))

m_sorted = m.sort_values('delta_valor', ascending=False)
for ri, row in enumerate(m_sorted.itertuples(index=False), start=4):
    cliente, ped_b, val, ult, k, val_s, ped_s, dv, dp = row
    ws4.cell(ri, 1, cliente)
    ws4.cell(ri, 2, int(ped_b))
    ws4.cell(ri, 3, int(ped_s))
    ws4.cell(ri, 4, val)
    ws4.cell(ri, 5, val_s)
    ws4.cell(ri, 6, int(dp))
    ws4.cell(ri, 7, dv)
    for ci in range(1, 8):
        ws4.cell(ri, ci).font = ARIAL
        ws4.cell(ri, ci).border = BORDER
        if ci > 1:
            ws4.cell(ri, ci).alignment = RIGHT
        if ci in (4, 5, 7):
            ws4.cell(ri, ci).number_format = '#,##0.00'
    if dv > 100:
        ws4.cell(ri, 7).fill = FILL_OK
    elif dv < -100:
        ws4.cell(ri, 7).fill = FILL_DANGER

ws4.column_dimensions['A'].width = 55
for col_letter in 'BCDEFG':
    ws4.column_dimensions[col_letter].width = 15

# Aba 5 - Naturezas excluidas
ws5 = wb.create_sheet('Naturezas Excluidas')
ws5.merge_cells('A1:G1')
ws5['A1'] = f'NATUREZAS EXCLUIDAS PELO FILTRO financial_flag=F - SAC - R$ {df_nat["faturamento"].sum():,.2f}'
ws5['A1'].font = ARIAL_TITLE
ws5['A1'].fill = FILL_PRIMARY
ws5['A1'].alignment = CENTER
ws5.row_dimensions[1].height = 30

ws5.merge_cells('A2:G2')
ws5['A2'] = 'ACAO: marcar coluna DECISAO ALVES/DIEGO como INCLUIR / EXCLUIR / DUVIDOSO'
ws5['A2'].font = ARIAL_BOLD
ws5['A2'].fill = FILL_WARN
ws5['A2'].alignment = CENTER

hdrs = ['nature_code', 'Nome da natureza', 'Flag', 'Clientes', 'Pedidos', 'Faturamento (R$)', 'DECISAO ALVES/DIEGO']
for i, h in enumerate(hdrs):
    style_hdr(ws5.cell(4, i + 1, h))

for ri, row in enumerate(df_nat.itertuples(index=False), start=5):
    nc, nm, flag, cli, ped, fat = row
    ws5.cell(ri, 1, nc)
    ws5.cell(ri, 2, nm)
    ws5.cell(ri, 3, flag)
    ws5.cell(ri, 4, int(cli))
    ws5.cell(ri, 5, int(ped))
    ws5.cell(ri, 6, fat)
    ws5.cell(ri, 6).number_format = '#,##0.00'
    ws5.cell(ri, 7, '')
    for ci in range(1, 8):
        ws5.cell(ri, ci).font = ARIAL
        ws5.cell(ri, ci).border = BORDER
        if ci in (4, 5, 6):
            ws5.cell(ri, ci).alignment = RIGHT
    if 'VENDA' in str(nm):
        for ci in range(1, 8):
            ws5.cell(ri, ci).fill = FILL_DANGER
        ws5.cell(ri, 7, 'SUSPEITO - nome diz VENDA mas flag != F')

ws5.column_dimensions['A'].width = 13
ws5.column_dimensions['B'].width = 52
ws5.column_dimensions['C'].width = 8
ws5.column_dimensions['D'].width = 10
ws5.column_dimensions['E'].width = 10
ws5.column_dimensions['F'].width = 18
ws5.column_dimensions['G'].width = 45

out = r'C:\Users\gusta\Downloads\Auditoria_RFV_SAC_Abril2026.xlsx'
wb.save(out)
print(f'OK: {out}')
print(f'Naturezas excluidas SAC: {len(df_nat)} | R$ {df_nat["faturamento"].sum():,.2f}')
print(f'Duplicatas detectadas: {len(df_dups)}')
