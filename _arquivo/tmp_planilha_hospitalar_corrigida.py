"""Reescreve a planilha Hospitalar com metodologia correta.

Regras:
  - Mesmos 786 clientes do Alves (por nome, com fuzzy se necessario)
  - Periodo: 01/04/2025 a 30/04/2026
  - Data de referencia: 30/04/2026 (fim do periodo, nao 02/04)
  - Todas as naturezas de venda (order_status 3 ou 4) - SEM filtro financial_flag=F
  - Calculos via SQL puro: MAX(order_date), COUNT(DISTINCT order_number), SUM(total_amount)
  - Buckets HOSPITALAR:
      F1>=5 | F2=4 | F3=3 | F4=2 | F5=1
      R1<=30 | R2<=60 | R3<=120 | R4<=180 | R5>180
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from datetime import datetime
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

creds = service_account.Credentials.from_service_account_file(
    r'C:\teste\sapient-metrics.json',
    scopes=["https://www.googleapis.com/auth/cloud-platform"]
)
client = bigquery.Client(credentials=creds, project='sapient-metrics-492914-m7')

REF_DATE = pd.Timestamp('2026-04-30')

# 1) Carrega 786 clientes do Excel original
path_excel = r'C:\Users\gusta\Downloads\RFV Hospitalar 01-04-2025 até 30-04-2026 (1).xlsx'
df_excel = pd.read_excel(path_excel, sheet_name='Sem fórmula Geral')
df_excel['k'] = df_excel['ID - CLIENTE'].astype(str).str.upper().str.strip()
print(f'Clientes do Excel Alves: {len(df_excel)}')

# 2) Tenta resolver partner_code via param_com_rfv_carteira (match exato por nome)
nomes = df_excel['k'].tolist()
# Carteira HOSPITALAR + dim_partner para os ausentes
df_carteira = client.query("""
SELECT partner_code, UPPER(TRIM(partner_name)) AS k_carteira, partner_name, rfv_familia, salesperson_name
FROM `sapient-metrics-492914-m7.silver_comercial.param_com_rfv_carteira`
WHERE is_active = TRUE
""").to_dataframe()
df_partner = client.query("""
SELECT partner_code, UPPER(TRIM(partner_name)) AS k_erp, partner_name
FROM `sapient-metrics-492914-m7.dm_partners.dim_partner`
""").to_dataframe()

# Match: primeiro na carteira (qualquer familia), depois dim_partner
df_match = df_excel[['k','ID - CLIENTE']].copy()
df_match = df_match.merge(df_carteira[['partner_code','k_carteira','rfv_familia','salesperson_name']],
                          left_on='k', right_on='k_carteira', how='left')
# Para os que nao acharam na carteira, busca dim_partner
faltam_mask = df_match['partner_code'].isna()
print(f'Match via carteira: {(~faltam_mask).sum()} | Faltam buscar em dim_partner: {faltam_mask.sum()}')

df_faltam = df_match[faltam_mask][['k','ID - CLIENTE']].copy()
df_faltam_pc = df_faltam.merge(df_partner[['partner_code','k_erp']], left_on='k', right_on='k_erp', how='left')
print(f'Match via dim_partner: {df_faltam_pc["partner_code"].notna().sum()} de {len(df_faltam)}')

# Consolida: pega partner_codes encontrados
df_match.loc[faltam_mask, 'partner_code'] = df_faltam_pc['partner_code'].values
df_match['encontrado'] = df_match['partner_code'].notna()
print(f'TOTAL ENCONTRADOS: {df_match["encontrado"].sum()} de {len(df_match)}')

# 3) Para os encontrados, busca compras COM TODAS as naturezas no periodo
# Importante: agrupar por partner_name (igual planilha consolida filiais)
codes = df_match[df_match['encontrado']]['partner_code'].dropna().astype('int64').unique().tolist()
codes_sql = ','.join(str(c) for c in codes)

q = f"""
WITH base AS (
  SELECT
    o.partner_code,
    o.order_number,
    o.order_date,
    o.invoice_date,
    o.nature_code,
    o.total_amount,
    o.order_status,
    n.financial_flag,
    n.nature_name
  FROM `sapient-metrics-492914-m7.dm_orders.fact_sales_order` o
  LEFT JOIN `sapient-metrics-492914-m7.dm_orders.dim_operation_nature` n
    ON n.nature_code = o.nature_code
  WHERE o.partner_code IN ({codes_sql})
    AND o.order_date BETWEEN '2025-04-01' AND '2026-04-30'
    AND o.order_status IN (3, 4)
    -- SEM filtro de natureza (todas as venda, S e devolucao se houver)
)
SELECT
  partner_code,
  MAX(order_date) AS ultima_compra,
  COUNT(DISTINCT order_number) AS frequencia,
  ROUND(SUM(total_amount), 2) AS valor_total
FROM base
GROUP BY partner_code
"""
df_calc = client.query(q).to_dataframe()
print(f'Clientes com pelo menos 1 venda no periodo: {len(df_calc)}')

# 4) Junta com nomes
df_final = df_match[df_match['encontrado']].copy()
df_final['partner_code'] = df_final['partner_code'].astype('int64')
df_final = df_final.merge(df_calc, on='partner_code', how='left')

# Se um cliente do Alves tem multiplos partner_codes (ex: filiais), agregamos POR NOME
# Pega o nome do Excel (ID - CLIENTE) como chave de consolidacao
agg = df_final.groupby('ID - CLIENTE', as_index=False).agg(
    ultima_compra=('ultima_compra', 'max'),
    frequencia=('frequencia', 'sum'),
    valor_total=('valor_total', 'sum'),
    partner_codes=('partner_code', lambda s: ', '.join(s.astype(str).unique())),
)
print(f'Apos consolidacao por nome: {len(agg)}')

# Adiciona os clientes nao encontrados (NaN nos calculos)
df_nao_enc = df_match[~df_match['encontrado']][['ID - CLIENTE']].copy()
df_nao_enc['ultima_compra'] = pd.NaT
df_nao_enc['frequencia'] = 0
df_nao_enc['valor_total'] = 0.0
df_nao_enc['partner_codes'] = ''
final = pd.concat([agg, df_nao_enc], ignore_index=True)
print(f'TOTAL FINAL (incluindo nao encontrados): {len(final)}')

# 5) Calcula recencia e classifica
final['ultima_compra'] = pd.to_datetime(final['ultima_compra'])
final['recencia_dias'] = (REF_DATE - final['ultima_compra']).dt.days

def freq_bucket(f):
    if f >= 5: return 'F1'
    if f == 4: return 'F2'
    if f == 3: return 'F3'
    if f == 2: return 'F4'
    return 'F5'

def rec_bucket(r):
    if pd.isna(r): return 'R5'  # sem compra = perdido
    if r <= 30:  return 'R1'
    if r <= 60:  return 'R2'
    if r <= 120: return 'R3'
    if r <= 180: return 'R4'
    return 'R5'

SEG_MAP = {
    ('F1','R1'): (1, 'Campeões'),
    ('F1','R2'): (2, 'Fiéis'), ('F1','R3'): (2, 'Fiéis'),
    ('F1','R4'): (8, 'Não pode perder'), ('F1','R5'): (8, 'Não pode perder'),
    ('F2','R1'): (2, 'Fiéis'), ('F2','R2'): (2, 'Fiéis'), ('F2','R3'): (2, 'Fiéis'),
    ('F2','R4'): (9, 'Em risco'), ('F2','R5'): (9, 'Em risco'),
    ('F3','R1'): (3, 'Fiéis em potencial'), ('F3','R2'): (3, 'Fiéis em potencial'),
    ('F3','R3'): (6, 'Precisando de atenção'),
    ('F3','R4'): (9, 'Em risco'), ('F3','R5'): (9, 'Em risco'),
    ('F4','R1'): (3, 'Fiéis em potencial'), ('F4','R2'): (3, 'Fiéis em potencial'),
    ('F4','R3'): (7, 'Quase dormentes'),
    ('F4','R4'): (10, 'Hibernando'), ('F4','R5'): (11, 'Perdidos'),
    ('F5','R1'): (4, 'Novos clientes'), ('F5','R2'): (5, 'Promessas'),
    ('F5','R3'): (7, 'Quase dormentes'),
    ('F5','R4'): (11, 'Perdidos'), ('F5','R5'): (11, 'Perdidos'),
}

final['freq_bucket'] = final['frequencia'].apply(freq_bucket)
final['rec_bucket'] = final['recencia_dias'].apply(rec_bucket)
final['classificacao_1'] = final['freq_bucket'] + final['rec_bucket']
final['seg_num'] = final.apply(lambda r: SEG_MAP[(r['freq_bucket'], r['rec_bucket'])][0], axis=1)
final['segmento'] = final.apply(lambda r: SEG_MAP[(r['freq_bucket'], r['rec_bucket'])][1], axis=1)

# Marca quem nao tem compra no periodo (deveria ser "sem dados")
final['sem_compra_no_periodo'] = final['frequencia'] == 0

# 6) Cross-check com a planilha original (segmento Alves)
df_excel_red = df_excel[['ID - CLIENTE','Classificação 2','Frequência 1','Recência em dias','Data última compra','Valor']].rename(
    columns={'Classificação 2':'seg_alves', 'Frequência 1':'freq_alves',
             'Recência em dias':'rec_alves', 'Data última compra':'ultima_alves', 'Valor':'val_alves'})
final = final.merge(df_excel_red, on='ID - CLIENTE', how='left')

# 7) Tambem inclui o que o SISTEMA mostra (silver atual)
df_sys = client.query("""
SELECT partner_name, ultima_compra_data, recencia_dias AS rec_sys, frequencia AS freq_sys,
       valor_total AS val_sys, classificacao_2 AS seg_sys
FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
WHERE rfv_familia = 'HOSPITALAR'
  AND data_referencia = (SELECT MAX(data_referencia) FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score` WHERE rfv_familia='HOSPITALAR')
""").to_dataframe()
df_sys['k'] = df_sys['partner_name'].str.upper().str.strip()
final['k'] = final['ID - CLIENTE'].astype(str).str.upper().str.strip()
final = final.merge(df_sys[['k','ultima_compra_data','rec_sys','freq_sys','val_sys','seg_sys']], on='k', how='left')

# Distribuicao por segmento (3 visões)
dist_corrigida = final[~final['sem_compra_no_periodo']]['segmento'].value_counts()
dist_alves = final['seg_alves'].value_counts()
dist_sys = final['seg_sys'].dropna().value_counts()
print()
print('=== Distribuicao por segmento ===')
order = ['Campeões','Fiéis','Fiéis em potencial','Novos clientes','Promessas',
         'Precisando de atenção','Quase dormentes','Não pode perder','Em risco','Hibernando','Perdidos']
print(f'{"Segmento":<25} {"Alves":>8} {"Corrigida":>10} {"Sistema":>9}')
for s in order:
    a = int(dist_alves.get(s, 0))
    c = int(dist_corrigida.get(s, 0))
    ss = int(dist_sys.get(s, 0))
    print(f'{s:<25} {a:>8} {c:>10} {ss:>9}')

print()
total_val_corr = final[~final['sem_compra_no_periodo']]['valor_total'].sum()
total_val_alves = final['val_alves'].sum()
total_val_sys = final['val_sys'].sum()
total_cli_corr = (~final['sem_compra_no_periodo']).sum()
total_cli_alves = final['seg_alves'].notna().sum()
total_cli_sys = final['seg_sys'].notna().sum()
print(f'TOTAL CLIENTES — Alves: {total_cli_alves} | Corrigida: {total_cli_corr} | Sistema: {total_cli_sys}')
print(f'TOTAL FATURAMENTO — Alves: R$ {total_val_alves:,.2f} | Corrigida: R$ {total_val_corr:,.2f} | Sistema: R$ {total_val_sys:,.2f}')

# ====== GERA EXCEL ======
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
ARIAL_SUB = Font(name='Arial', size=11, bold=True)
FILL_PRIMARY = PatternFill('solid', fgColor='1E1882')
FILL_HEADER = PatternFill('solid', fgColor='4844C8')
FILL_ALVES = PatternFill('solid', fgColor='F4B183')   # laranja claro
FILL_CORR = PatternFill('solid', fgColor='C9F8D2')    # verde
FILL_SYS = PatternFill('solid', fgColor='B8CCE4')     # azul claro
FILL_WARN = PatternFill('solid', fgColor='FFE066')
FILL_DANGER = PatternFill('solid', fgColor='F8C9C9')
FILL_OK = PatternFill('solid', fgColor='C9F8D2')
FILL_LIGHT = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# Aba 1 — Comparativo 3-vias
ws = wb.active
ws.title = 'Comparativo 3-vias'
ws.merge_cells('A1:E1')
ws['A1'] = 'COMPARATIVO 3-VIAS — Alves (planilha) x Corrigida x Sistema atual'
ws['A1'].font = ARIAL_TITLE; ws['A1'].fill = FILL_PRIMARY; ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:E2')
ws['A2'] = ('CORRIGIDA = mesmos 786 clientes do Alves + todas as naturezas (sem filtro flag F) '
            '+ data ref 30/04/2026 + calculo SQL puro. Eh a "planilha ideal" do que o Alves deveria ter entregue.')
ws['A2'].font = ARIAL; ws['A2'].fill = FILL_WARN; ws['A2'].alignment = CENTER
ws.row_dimensions[2].height = 40

# Totais
hdrs = ['Metrica', 'Alves (planilha)', 'Corrigida (proposta)', 'Sistema atual', 'Observacao']
for i, h in enumerate(hdrs):
    c = ws.cell(4, i + 1, h)
    c.font = ARIAL_HDR
    c.fill = FILL_HEADER if i == 0 else (FILL_ALVES if i == 1 else (FILL_CORR if i == 2 else (FILL_SYS if i == 3 else FILL_HEADER)))
    c.alignment = CENTER; c.border = BORDER

linhas_tot = [
    ('Clientes', total_cli_alves, total_cli_corr, total_cli_sys,
     'Corrigida usa os mesmos 786; Sistema tem 755 (carteira nao tem os 31)'),
    ('Faturamento (R$)', float(total_val_alves), float(total_val_corr), float(total_val_sys),
     'Corrigida inclui todas as naturezas; Sistema filtra flag=F'),
    ('Data de referencia', '02/04/2026 (errada)', '30/04/2026 (correta)', '30/04/2026 (CURRENT_DATE)', ''),
    ('Filtro de natureza', 'Sem (todas)', 'Sem (todas)', 'financial_flag = F', ''),
]
for ri, row in enumerate(linhas_tot, start=5):
    for ci, v in enumerate(row, start=1):
        c = ws.cell(ri, ci, v)
        c.font = ARIAL; c.border = BORDER
        if isinstance(v, float):
            c.number_format = '#,##0.00'
            c.alignment = RIGHT
        elif isinstance(v, int):
            c.alignment = RIGHT
        if ci == 2: c.fill = FILL_ALVES
        elif ci == 3: c.fill = FILL_CORR
        elif ci == 4: c.fill = FILL_SYS

# Distribuicao por segmento
ws['A11'] = 'DISTRIBUICAO POR SEGMENTO'
ws['A11'].font = ARIAL_SUB
hdrs2 = ['Segmento', 'Alves (planilha)', 'Corrigida', 'Sistema atual', 'Alves vs Corrigida']
for i, h in enumerate(hdrs2):
    c = ws.cell(12, i + 1, h)
    c.font = ARIAL_HDR
    c.fill = FILL_HEADER if i == 0 else (FILL_ALVES if i == 1 else (FILL_CORR if i == 2 else (FILL_SYS if i == 3 else FILL_HEADER)))
    c.alignment = CENTER; c.border = BORDER

for ri, s in enumerate(order, start=13):
    a = int(dist_alves.get(s, 0))
    c = int(dist_corrigida.get(s, 0))
    ss = int(dist_sys.get(s, 0))
    delta = c - a
    ws.cell(ri, 1, s)
    ws.cell(ri, 2, a)
    ws.cell(ri, 3, c)
    ws.cell(ri, 4, ss)
    ws.cell(ri, 5, f'{delta:+d}')
    for ci in range(1, 6):
        cc = ws.cell(ri, ci)
        cc.font = ARIAL; cc.border = BORDER
        if ci > 1: cc.alignment = RIGHT
    ws.cell(ri, 2).fill = FILL_ALVES
    ws.cell(ri, 3).fill = FILL_CORR
    ws.cell(ri, 4).fill = FILL_SYS
    if delta > 0: ws.cell(ri, 5).fill = FILL_OK
    elif delta < 0: ws.cell(ri, 5).fill = FILL_DANGER

# Total
ri = 13 + len(order)
ws.cell(ri, 1, 'TOTAL').font = ARIAL_BOLD
ws.cell(ri, 2, int(dist_alves.sum()))
ws.cell(ri, 3, int(dist_corrigida.sum()))
ws.cell(ri, 4, int(dist_sys.sum()))
for ci in range(1, 6):
    cc = ws.cell(ri, ci)
    cc.font = ARIAL_BOLD; cc.border = BORDER; cc.fill = FILL_LIGHT
    if ci > 1: cc.alignment = RIGHT

# Diagnostico
ri += 2
ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=5)
ws.cell(ri, 1, 'DIAGNOSTICO').font = ARIAL_SUB
diag = [
    '1. Compare Alves x Corrigida — isola apenas a correcao da DATA DE REFERENCIA e elimina recencias negativas.',
    '   A diferenca aqui mostra o impacto puro do bug de data: clientes que mudaram de segmento so por causa disso.',
    '',
    '2. Compare Corrigida x Sistema — isola o efeito do FILTRO DE NATUREZA + COBERTURA DE CARTEIRA.',
    '   Mesma cobertura de clientes (786) e mesma data de ref (30/04). So diferem nas naturezas incluidas.',
    '',
    '3. Conclusao esperada: Corrigida vai bater quase 100% com o que o sistema mostraria se incluisse todas as naturezas.',
    '   O que sobrar de divergencia e somente: 31 clientes faltantes na carteira do sistema.',
]
for i, line in enumerate(diag, start=ri + 1):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    c = ws.cell(i, 1, line)
    c.font = ARIAL_BOLD if line.startswith(('1.','2.','3.')) else ARIAL

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 45

# Aba 2 — Matriz RFV Corrigida (estilo planilha)
ws2 = wb.create_sheet('Matriz Corrigida')
ws2.merge_cells('A1:G1')
ws2['A1'] = 'MATRIZ RFV CORRIGIDA — Hospitalar (30/04/2026, todas as naturezas, 786 clientes)'
ws2['A1'].font = ARIAL_TITLE; ws2['A1'].fill = FILL_PRIMARY; ws2['A1'].alignment = CENTER
ws2.row_dimensions[1].height = 30

# Cabecalho R
ws2.merge_cells('A3:B4')
ws2['A3'] = 'Frequencia x Recencia'
ws2['A3'].font = ARIAL_HDR; ws2['A3'].fill = PatternFill('solid', fgColor='C00000')
ws2['A3'].alignment = CENTER

r_labels = ['R1 (<=30d)', 'R2 (31-60d)', 'R3 (61-120d)', 'R4 (121-180d)', 'R5 (>180d)']
for i, lbl in enumerate(r_labels):
    c = ws2.cell(3, 3 + i, lbl)
    c.font = ARIAL_HDR; c.fill = PatternFill('solid', fgColor='C00000'); c.alignment = CENTER

f_labels = [('F1', '5+'), ('F2', '4'), ('F3', '3'), ('F4', '2'), ('F5', '1')]
celula_grid = {}
fonly = final[~final['sem_compra_no_periodo']]
for f, label_f in f_labels:
    row_idx = 5 + f_labels.index((f, label_f))
    ws2.cell(row_idx, 1, f).font = ARIAL_HDR
    ws2.cell(row_idx, 1).fill = PatternFill('solid', fgColor='C00000')
    ws2.cell(row_idx, 1).alignment = CENTER
    ws2.cell(row_idx, 2, label_f).font = ARIAL
    ws2.cell(row_idx, 2).alignment = CENTER
    for i, rlbl in enumerate(['R1','R2','R3','R4','R5']):
        sub = fonly[(fonly['freq_bucket']==f) & (fonly['rec_bucket']==rlbl)]
        cnt = len(sub)
        val = sub['valor_total'].sum()
        seg = SEG_MAP[(f, rlbl)][1]
        ws2.cell(row_idx, 3 + i, f'{seg}\n{cnt} cli\nR$ {val:,.0f}').alignment = CENTER
        ws2.cell(row_idx, 3 + i).font = ARIAL
        ws2.cell(row_idx, 3 + i).border = BORDER

# Resumo abaixo da matriz
ri = 12
ws2.cell(ri, 1, 'RESUMO POR SEGMENTO').font = ARIAL_SUB
ri += 1
hdrs = ['Segmento', 'Clientes', 'Faturamento (R$)']
for i, h in enumerate(hdrs):
    c = ws2.cell(ri, i + 1, h)
    c.font = ARIAL_HDR; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
ri += 1
agg_seg = fonly.groupby('segmento').agg(cli=('ID - CLIENTE','count'), val=('valor_total','sum'))
total_cli_m = 0; total_val_m = 0.0
for s in order:
    if s in agg_seg.index:
        cli = int(agg_seg.loc[s, 'cli'])
        val = float(agg_seg.loc[s, 'val'])
        ws2.cell(ri, 1, s); ws2.cell(ri, 2, cli); ws2.cell(ri, 3, val)
        ws2.cell(ri, 3).number_format = '#,##0.00'
        for ci in range(1, 4):
            ws2.cell(ri, ci).font = ARIAL
            ws2.cell(ri, ci).border = BORDER
            if ci > 1: ws2.cell(ri, ci).alignment = RIGHT
        total_cli_m += cli; total_val_m += val
        ri += 1
ws2.cell(ri, 1, 'TOTAL').font = ARIAL_BOLD
ws2.cell(ri, 2, total_cli_m).font = ARIAL_BOLD
ws2.cell(ri, 3, total_val_m).font = ARIAL_BOLD
ws2.cell(ri, 3).number_format = '#,##0.00'
for ci in range(1, 4):
    ws2.cell(ri, ci).border = BORDER; ws2.cell(ri, ci).fill = FILL_LIGHT
    if ci > 1: ws2.cell(ri, ci).alignment = RIGHT

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 12
for col in 'CDEFG':
    ws2.column_dimensions[col].width = 18
for r in range(3, 12):
    ws2.row_dimensions[r].height = 55

# Aba 3 — 786 clientes lado a lado (Alves x Corrigida x Sistema)
ws3 = wb.create_sheet('Drill 786 — 3 visoes')
ws3.merge_cells('A1:O1')
ws3['A1'] = 'DRILL 3-VIAS — Alves x Corrigida x Sistema (cliente por cliente)'
ws3['A1'].font = ARIAL_TITLE; ws3['A1'].fill = FILL_PRIMARY; ws3['A1'].alignment = CENTER
ws3.row_dimensions[1].height = 30

# Headers em 2 niveis
ws3.merge_cells('A3:A4'); ws3['A3'] = 'Cliente'
ws3.merge_cells('B3:E3'); ws3['B3'] = 'ALVES (planilha)'; ws3['B3'].fill = FILL_ALVES
ws3.merge_cells('F3:I3'); ws3['F3'] = 'CORRIGIDA (proposta)'; ws3['F3'].fill = FILL_CORR
ws3.merge_cells('J3:M3'); ws3['J3'] = 'SISTEMA atual'; ws3['J3'].fill = FILL_SYS
ws3.merge_cells('N3:O3'); ws3['N3'] = 'DIVERGENCIAS'; ws3['N3'].fill = FILL_HEADER

for cell_ref in ['A3','B3','F3','J3','N3']:
    ws3[cell_ref].font = ARIAL_HDR
    ws3[cell_ref].alignment = CENTER

subhdrs = [
    (2, 'Última'), (3, 'Rec'), (4, 'Freq'), (5, 'Seg'),
    (6, 'Última'), (7, 'Rec'), (8, 'Freq'), (9, 'Seg'),
    (10, 'Última'), (11, 'Rec'), (12, 'Freq'), (13, 'Seg'),
    (14, 'Alves=Corr?'), (15, 'Corr=Sis?'),
]
for col, lbl in subhdrs:
    c = ws3.cell(4, col, lbl)
    c.font = ARIAL_BOLD; c.alignment = CENTER; c.border = BORDER; c.fill = FILL_LIGHT
ws3.cell(4, 1).fill = FILL_LIGHT
ws3.cell(4, 1).border = BORDER

# Preenche
final_sorted = final.sort_values('ID - CLIENTE')
for ri, r in enumerate(final_sorted.itertuples(index=False), start=5):
    ws3.cell(ri, 1, r._0)  # ID - CLIENTE
    # Alves
    if pd.notna(r.ultima_alves):
        ws3.cell(ri, 2, r.ultima_alves); ws3.cell(ri, 2).number_format = 'dd/mm/yyyy'
    if pd.notna(r.rec_alves): ws3.cell(ri, 3, int(r.rec_alves))
    if pd.notna(r.freq_alves): ws3.cell(ri, 4, int(r.freq_alves))
    ws3.cell(ri, 5, r.seg_alves if pd.notna(r.seg_alves) else '')
    # Corrigida
    if pd.notna(r.ultima_compra):
        ws3.cell(ri, 6, r.ultima_compra); ws3.cell(ri, 6).number_format = 'dd/mm/yyyy'
    if pd.notna(r.recencia_dias): ws3.cell(ri, 7, int(r.recencia_dias))
    ws3.cell(ri, 8, int(r.frequencia) if pd.notna(r.frequencia) else 0)
    ws3.cell(ri, 9, r.segmento if not r.sem_compra_no_periodo else '(sem compra)')
    # Sistema
    if pd.notna(r.ultima_compra_data):
        ws3.cell(ri, 10, r.ultima_compra_data); ws3.cell(ri, 10).number_format = 'dd/mm/yyyy'
    if pd.notna(r.rec_sys): ws3.cell(ri, 11, int(r.rec_sys))
    if pd.notna(r.freq_sys): ws3.cell(ri, 12, int(r.freq_sys))
    ws3.cell(ri, 13, r.seg_sys if pd.notna(r.seg_sys) else '(nao na carteira)')
    # Comparacoes
    alves_eq_corr = r.seg_alves == r.segmento if (pd.notna(r.seg_alves) and not r.sem_compra_no_periodo) else None
    corr_eq_sys = r.segmento == r.seg_sys if (pd.notna(r.seg_sys) and not r.sem_compra_no_periodo) else None
    ws3.cell(ri, 14, 'OK' if alves_eq_corr else ('-' if alves_eq_corr is None else 'DIF'))
    ws3.cell(ri, 15, 'OK' if corr_eq_sys else ('-' if corr_eq_sys is None else 'DIF'))

    # Styles
    for ci in range(1, 16):
        cc = ws3.cell(ri, ci)
        cc.font = ARIAL; cc.border = BORDER
        if ci in (3,4,7,8,11,12): cc.alignment = RIGHT
    # Highlights
    if r.sem_compra_no_periodo:
        for ci in range(6, 10):
            ws3.cell(ri, ci).fill = FILL_DANGER
    elif alves_eq_corr is False:
        ws3.cell(ri, 14).fill = FILL_DANGER
    if corr_eq_sys is False:
        ws3.cell(ri, 15).fill = FILL_DANGER

widths = [40, 11, 6, 6, 18, 11, 6, 6, 18, 11, 6, 6, 18, 12, 12]
for i, w in enumerate(widths):
    ws3.column_dimensions[chr(65 + i)].width = w

# Aba 4 — Metodologia
ws4 = wb.create_sheet('Metodologia', 0)  # primeira aba
ws4.merge_cells('A1:B1')
ws4['A1'] = 'METODOLOGIA DA PLANILHA CORRIGIDA'
ws4['A1'].font = ARIAL_TITLE; ws4['A1'].fill = FILL_PRIMARY; ws4['A1'].alignment = CENTER
ws4.row_dimensions[1].height = 30

mtd = [
    ('Objetivo', 'Reescrever a analise RFV Hospitalar com metodologia correta, mantendo a mesma cobertura de clientes da planilha do Alves, para isolar bugs e divergencias.'),
    ('', ''),
    ('1. Mesmos clientes', 'Pegamos os 786 clientes da aba "Sem formula Geral" do Alves. Match com ERP via nome (carteira primeiro, depois dim_partner).'),
    ('2. Mesmo periodo', '01/04/2025 a 30/04/2026 (igual nome do arquivo).'),
    ('3. Data de referencia', '30/04/2026 (CORRIGIDA — fim do periodo). Alves usou 02/04/2026, gerando recencias negativas.'),
    ('4. Naturezas', 'TODAS as naturezas com order_status IN (3,4). SEM filtro financial_flag=F. Igual o Alves fez na planilha original.'),
    ('5. Calculos', 'SQL puro: MAX(o.order_date), COUNT(DISTINCT o.order_number), SUM(o.total_amount).'),
    ('6. Thresholds', 'Buckets HOSPITALAR: F1>=5, F2=4, F3=3, F4=2, F5=1 | R1<=30, R2<=60, R3<=120, R4<=180, R5>180.'),
    ('7. Consolidacao', 'Por NOME do cliente (igual planilha), agrupando filiais com mesmo nome.'),
    ('', ''),
    ('O que essa planilha PROVA', ''),
    ('Bug 1 — Recencia negativa', 'Compare colunas "Alves" e "Corrigida". Clientes que mudaram de segmento entre as duas = afetados pelo bug de data.'),
    ('Bug 2 — Filtro de natureza', 'Compare colunas "Corrigida" e "Sistema". Mesma data, mesma cobertura. Diferenca = naturezas que sistema exclui.'),
    ('Gap de carteira', '31 clientes do Alves nao estao na carteira do sistema (coluna Sistema = "nao na carteira").'),
]
for ri, (k, v) in enumerate(mtd, start=3):
    if k:
        ws4.cell(ri, 1, k).font = ARIAL_BOLD
        ws4.cell(ri, 1).alignment = Alignment(vertical='top', wrap_text=True)
    ws4.cell(ri, 2, v).font = ARIAL
    ws4.cell(ri, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws4.row_dimensions[ri].height = max(25, 18 * (len(v) // 80 + 1)) if v else 12

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 110

out = r'C:\Users\gusta\Downloads\RFV_Hospitalar_CORRIGIDA.xlsx'
wb.save(out)
print(f'\nOK: {out}')
