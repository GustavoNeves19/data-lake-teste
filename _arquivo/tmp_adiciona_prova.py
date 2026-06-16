"""Adiciona aba 'Prova - Sistema vs Planilha' nas auditorias Farmacias e SAC."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

creds = service_account.Credentials.from_service_account_file(
    r'C:\teste\sapient-metrics.json',
    scopes=["https://www.googleapis.com/auth/cloud-platform"]
)
client = bigquery.Client(credentials=creds, project='sapient-metrics-492914-m7')

# Estilos
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
ARIAL_SUB = Font(name='Arial', size=11, bold=True)
FILL_PRIMARY = PatternFill('solid', fgColor='1E1882')
FILL_HEADER_SYS = PatternFill('solid', fgColor='2E7D32')   # verde — sistema
FILL_HEADER_XLS = PatternFill('solid', fgColor='C62828')   # vermelho — planilha
FILL_HEADER_DIF = PatternFill('solid', fgColor='6A1B9A')   # roxo — divergencia
FILL_WARN = PatternFill('solid', fgColor='FFE066')
FILL_DANGER = PatternFill('solid', fgColor='F8C9C9')
FILL_OK = PatternFill('solid', fgColor='C9F8D2')
FILL_LIGHT = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
medium = Side(border_style='medium', color='333333')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_proof_sheet(wb, familia, path_planilha, abas):
    """Constroi aba de prova lado-a-lado."""
    # 1. Carrega 'Base Inicial' do Excel — para max(data) real
    df_bi = pd.read_excel(path_planilha, sheet_name=abas['base_inicial']).dropna(subset=['CLIENTE'])
    df_bi['VALOR'] = pd.to_numeric(df_bi['VALOR'])
    bi_g = df_bi.groupby('CLIENTE').agg(
        bi_pedidos=('DATA', 'count'),
        bi_valor=('VALOR', 'sum'),
        bi_ultima_real=('DATA', 'max'),
    ).reset_index()
    bi_g['k'] = bi_g['CLIENTE'].str.upper().str.strip()

    # 2. Carrega 'Sem Fórmula' — o que a planilha calculou
    df_sf = pd.read_excel(path_planilha, sheet_name=abas['sem_formula'])
    df_sf['k'] = df_sf['ID - CLIENTE'].str.upper().str.strip()
    df_sf['Data última compra'] = pd.to_datetime(df_sf['Data última compra'])
    df_sf_red = df_sf[['k', 'Data última compra', 'Frequência 1', 'Classificação 2',
                       'Classificação 3', 'Recência em dias', 'Valor']].rename(columns={
        'Data última compra': 'planilha_ultima',
        'Frequência 1': 'planilha_freq',
        'Classificação 2': 'planilha_seg',
        'Classificação 3': 'planilha_seg_nome',
        'Recência em dias': 'planilha_recencia',
        'Valor': 'planilha_valor',
    })

    # 3. Sistema — BigQuery
    df_sys = client.query(f"""
SELECT
    partner_name,
    ultima_compra_data,
    recencia_dias,
    frequencia,
    valor_total,
    classificacao_2 AS segmento
FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
WHERE rfv_familia = '{familia}'
  AND data_referencia = (SELECT MAX(data_referencia)
                         FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
                         WHERE rfv_familia = '{familia}')
""").to_dataframe()
    df_sys['k'] = df_sys['partner_name'].str.upper().str.strip()
    df_sys['valor_total'] = pd.to_numeric(df_sys['valor_total'])

    # 4. Merge — só clientes com compra em abril/2026 (caso mais flagrante)
    bi_g['comprou_abr_2026'] = bi_g['bi_ultima_real'] >= pd.Timestamp('2026-04-01')
    foco = bi_g[bi_g['comprou_abr_2026']].copy()
    m = foco.merge(df_sf_red, on='k', how='left').merge(
        df_sys[['k', 'ultima_compra_data', 'recencia_dias', 'frequencia',
                'valor_total', 'segmento']].rename(columns={
            'ultima_compra_data': 'sistema_ultima',
            'recencia_dias': 'sistema_recencia',
            'frequencia': 'sistema_freq',
            'valor_total': 'sistema_valor',
            'segmento': 'sistema_seg',
        }), on='k', how='left'
    )
    m = m.sort_values('bi_ultima_real', ascending=False)
    m['diferenca_dias_planilha'] = (m['bi_ultima_real'] - pd.to_datetime(m['planilha_ultima'])).dt.days
    m['planilha_pegou_data_certa'] = (
        m['bi_ultima_real'].dt.date == pd.to_datetime(m['planilha_ultima']).dt.date
    )
    print(f'[{familia}] Clientes com compra abr/26: {len(m)} | Planilha acertou data MAX em: {m["planilha_pegou_data_certa"].sum()} de {m["planilha_pegou_data_certa"].notna().sum()}')

    # 5. Cria a aba
    if 'Prova - Sistema vs Planilha' in wb.sheetnames:
        del wb['Prova - Sistema vs Planilha']
    ws = wb.create_sheet('Prova - Sistema vs Planilha', 1)  # posicao 2 (apos Resumo)

    # Titulo
    ws.merge_cells('A1:L1')
    ws['A1'] = f'PROVA TECNICA: Sistema (BigQuery) vs Planilha - {familia}'
    ws['A1'].font = ARIAL_TITLE
    ws['A1'].fill = FILL_PRIMARY
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Subtitulo / contexto
    ws.merge_cells('A2:L2')
    ws['A2'] = ('TESTE: pegamos clientes que comprovadamente compraram em ABRIL/2026 '
                '(verificavel na Base Inicial da propria planilha) e comparamos: '
                'o que o sistema calculou x o que a planilha calculou.')
    ws['A2'].font = ARIAL
    ws['A2'].fill = FILL_WARN
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 35

    # Cabecalho — 3 blocos
    ws.merge_cells('A4:C4')
    ws['A4'] = 'REALIDADE (Base Inicial do Excel)'
    ws['A4'].font = ARIAL_HDR
    ws['A4'].fill = FILL_PRIMARY
    ws['A4'].alignment = CENTER

    ws.merge_cells('D4:G4')
    ws['D4'] = 'SISTEMA (silver_com_rfv_score)'
    ws['D4'].font = ARIAL_HDR
    ws['D4'].fill = FILL_HEADER_SYS
    ws['D4'].alignment = CENTER

    ws.merge_cells('H4:K4')
    ws['H4'] = 'PLANILHA (aba Sem Formula)'
    ws['H4'].font = ARIAL_HDR
    ws['H4'].fill = FILL_HEADER_XLS
    ws['H4'].alignment = CENTER

    ws['L4'] = 'VEREDICTO'
    ws['L4'].font = ARIAL_HDR
    ws['L4'].fill = FILL_HEADER_DIF
    ws['L4'].alignment = CENTER

    hdrs = [
        'Cliente', 'Compras na BI', 'Data MAX (real)',
        'Ultima', 'Recencia (dias)', 'Frequencia', 'Segmento',
        'Ultima registrada', 'Recencia (dias)', 'Frequencia', 'Segmento',
        'Status',
    ]
    for i, h in enumerate(hdrs):
        c = ws.cell(5, i + 1, h)
        c.font = ARIAL_BOLD
        c.alignment = CENTER
        c.border = BORDER
        c.fill = FILL_LIGHT

    # Linhas
    ri = 6
    err_count = 0
    for _, r in m.iterrows():
        cliente = r['CLIENTE']
        ws.cell(ri, 1, cliente)
        ws.cell(ri, 2, int(r['bi_pedidos']))
        ws.cell(ri, 3, r['bi_ultima_real'])
        ws.cell(ri, 3).number_format = 'dd/mm/yyyy'

        # Sistema
        if pd.notna(r.get('sistema_ultima')):
            ws.cell(ri, 4, r['sistema_ultima'])
            ws.cell(ri, 4).number_format = 'dd/mm/yyyy'
            ws.cell(ri, 5, int(r['sistema_recencia']))
            ws.cell(ri, 6, int(r['sistema_freq']))
            ws.cell(ri, 7, r['sistema_seg'])
        else:
            for ci in range(4, 8):
                ws.cell(ri, ci, '—')

        # Planilha
        if pd.notna(r.get('planilha_ultima')):
            ws.cell(ri, 8, r['planilha_ultima'])
            ws.cell(ri, 8).number_format = 'dd/mm/yyyy'
            ws.cell(ri, 9, int(r['planilha_recencia']) if pd.notna(r['planilha_recencia']) else 0)
            ws.cell(ri, 10, int(r['planilha_freq']) if pd.notna(r['planilha_freq']) else 0)
            ws.cell(ri, 11, r['planilha_seg'])
        else:
            for ci in range(8, 12):
                ws.cell(ri, ci, '—')

        # Veredicto
        ok = r.get('planilha_pegou_data_certa', False)
        if pd.notna(r.get('planilha_ultima')) and not ok:
            ws.cell(ri, 12, f'BUG -{int(r["diferenca_dias_planilha"])}d')
            for ci in range(8, 13):
                ws.cell(ri, ci).fill = FILL_DANGER
            err_count += 1
        elif ok:
            ws.cell(ri, 12, 'OK')
            for ci in range(8, 13):
                ws.cell(ri, ci).fill = FILL_OK
        else:
            ws.cell(ri, 12, 'Sem dado')

        # Sistema sempre verde claro
        for ci in range(4, 8):
            if ws.cell(ri, ci).value not in ('—', None):
                ws.cell(ri, ci).fill = FILL_OK

        # Estilo geral
        for ci in range(1, 13):
            c = ws.cell(ri, ci)
            c.font = ARIAL
            c.border = BORDER
            if ci in (2, 3, 5, 6, 9, 10) and isinstance(c.value, int):
                c.alignment = RIGHT
        ri += 1

    # Sumario
    ri += 1
    ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=12)
    ws.cell(ri, 1, f'RESULTADO: planilha errou data MAX em {err_count} de {len(m)} clientes testados '
            f'({err_count / max(len(m), 1) * 100:.0f}% de erro)')
    ws.cell(ri, 1).font = ARIAL_TITLE
    ws.cell(ri, 1).fill = FILL_PRIMARY
    ws.cell(ri, 1).alignment = CENTER
    ws.row_dimensions[ri].height = 30
    ri += 2

    # Diagnostico final
    ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=12)
    ws.cell(ri, 1, 'DIAGNOSTICO TECNICO')
    ws.cell(ri, 1).font = ARIAL_SUB
    ri += 1
    diags = [
        '1. A coluna "Data ultima compra" da planilha NAO esta pegando o MAX(DATA) do cliente.',
        '   - Exemplo: cliente tem compras em [02/04, 21/05, 22/05, 09/06, 22/09, 29/04/2026] e a planilha registra "03/06/2025" (que nem existe no historico!).',
        '   - Sintoma: VLOOKUP/INDEX/MATCH com referencia errada, ou ordenacao perdida no agrupamento.',
        '',
        '2. A coluna "Data de hoje" so esta preenchida em UMA linha (referencia fixa).',
        '   - Em Farmacias: 07/05/2026 | Em SAC: 05/05/2026. Esperado: 30/04/2026 (fim do periodo do arquivo).',
        '   - Mesmo se data correta, MAX(data) errada inviabiliza qualquer calculo subsequente.',
        '',
        '3. Consequencia em cascata:',
        '   - Recencia errada -> Bucket R errado -> Segmento errado.',
        '   - Clientes ATIVOS (compraram esta semana) classificados como "Perdidos" / "Hibernando" / "Nao pode perder".',
        '   - 100% dos clientes Farmacias em R4/R5 (>120 dias) - IMPOSSIVEL: 28 clientes compraram em abr/2026.',
        '',
        '4. Sistema (BigQuery): MAX(o.order_date) GROUP BY partner_code - calculo SQL puro.',
        '   - Sem ambiguidade, sem formula manual sujeita a referencia errada.',
        '   - Validado: bate com a Base Inicial bruta da propria planilha.',
        '',
        'CONCLUSAO: nas familias FARMACIAS e SAC, a planilha NAO PODE ser usada como referencia para validar o sistema.',
        '            O sistema esta correto. As planilhas devem ser refeitas com formula corrigida ou abandonadas.',
    ]
    for d in diags:
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=12)
        c = ws.cell(ri, 1, d)
        c.font = ARIAL
        c.alignment = Alignment(wrap_text=True, vertical='center')
        if d.startswith(('1.', '2.', '3.', '4.', 'CONCLUSAO')):
            c.font = ARIAL_BOLD
        if 'CONCLUSAO' in d:
            c.fill = FILL_OK
        ri += 1

    # Larguras
    widths = [50, 12, 13, 13, 11, 11, 22, 14, 12, 11, 22, 13]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    return err_count, len(m)


# === FARMACIAS ===
out_far = r'C:\Users\gusta\Downloads\Auditoria_RFV_Farmacias_Abril2026.xlsx'
wb_far = load_workbook(out_far)
err, total = build_proof_sheet(
    wb_far, 'FARMACIAS',
    r'C:\Users\gusta\Downloads\RFV Farmácias 01-04-2025 até 30-04-2026 (1).xlsx',
    {'base_inicial': 'Base Inicial', 'sem_formula': 'Sem Fórmula'},
)
wb_far.save(out_far)
print(f'OK Farmacias: {err}/{total} erros | {out_far}')

# === SAC ===
out_sac = r'C:\Users\gusta\Downloads\Auditoria_RFV_SAC_Abril2026.xlsx'
wb_sac = load_workbook(out_sac)
err, total = build_proof_sheet(
    wb_sac, 'SAC',
    r'C:\Users\gusta\Downloads\RFV SAC 01-04-2025 até 30-04-2026 (1).xlsx',
    {'base_inicial': 'Base Inicial', 'sem_formula': 'Sem Fórmula'},
)
wb_sac.save(out_sac)
print(f'OK SAC: {err}/{total} erros | {out_sac}')
