"""Auditoria RFV Farmacias para reuniao Alves/Diego."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

creds = service_account.Credentials.from_service_account_file(
    r'C:\teste\sapient-metrics.json',
    scopes=["https://www.googleapis.com/auth/cloud-platform"]
)
client = bigquery.Client(credentials=creds, project='sapient-metrics-492914-m7')

# === Excel ===
path = r'C:\Users\gusta\Downloads\RFV Farmácias 01-04-2025 até 30-04-2026 (1).xlsx'
df_e = pd.read_excel(path, sheet_name='Base Inicial').dropna(subset=['CLIENTE'])
df_e['VALOR'] = pd.to_numeric(df_e['VALOR'])
ex_g = df_e.groupby('CLIENTE').agg(
    pedidos_brutos=('DATA', 'count'),
    valor=('VALOR', 'sum'),
    ultima=('DATA', 'max'),
).reset_index()
ex_g['k'] = ex_g['CLIENTE'].str.upper().str.strip()

# Sistema atual
df_s = client.query("""
SELECT c.partner_name AS cliente, COUNT(DISTINCT o.order_number) AS pedidos,
       ROUND(SUM(o.total_amount), 2) AS valor
FROM `sapient-metrics-492914-m7.dm_orders.fact_sales_order` o
JOIN `sapient-metrics-492914-m7.silver_comercial.param_com_rfv_carteira` c
  ON c.partner_code = o.partner_code AND c.is_active = TRUE
  AND c.salesperson_name NOT IN ('Eduardo', 'Karina')
JOIN `sapient-metrics-492914-m7.dm_orders.dim_operation_nature` n
  ON n.nature_code = o.nature_code AND n.financial_flag = 'F'
WHERE o.order_status IN (3, 4) AND c.rfv_familia = 'FARMACIAS'
  AND o.order_date BETWEEN '2025-04-01' AND '2026-04-30'
GROUP BY 1
""").to_dataframe()
df_s['k'] = df_s['cliente'].str.upper().str.strip()
df_s['valor'] = pd.to_numeric(df_s['valor'])

# Faltantes (so no Excel)
so_excel = ex_g[~ex_g['k'].isin(df_s['k'])].copy()

# Drill clientes em comum
m = ex_g.merge(df_s[['k', 'valor', 'pedidos']].rename(
    columns={'valor': 'valor_s', 'pedidos': 'pedidos_s'}), on='k', how='inner')
m['delta_valor'] = m['valor_s'] - m['valor']
m['delta_pedidos'] = m['pedidos_s'] - m['pedidos_brutos']

# Detecta duplicatas potenciais (mesma data + mesmo valor)
df_e_sorted = df_e.sort_values(['CLIENTE', 'DATA', 'VALOR']).reset_index(drop=True)
df_e_sorted['dup'] = df_e_sorted.duplicated(subset=['CLIENTE', 'DATA', 'VALOR'], keep=False)
df_dups = df_e_sorted[df_e_sorted['dup']].copy()
print(f'Excel: {len(df_e)} linhas, {len(df_dups)} potencialmente duplicadas')

# Quebra sistema por segmento
df_seg_sys = client.query("""
SELECT classificacao_3 AS seg_num, ANY_VALUE(classificacao_2) AS segmento,
       COUNT(DISTINCT partner_name) AS clientes,
       ROUND(SUM(valor_total), 2) AS faturamento
FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
WHERE rfv_familia = 'FARMACIAS'
  AND data_referencia = (SELECT MAX(data_referencia)
                         FROM `sapient-metrics-492914-m7.silver_comercial.silver_com_rfv_score`
                         WHERE rfv_familia = 'FARMACIAS')
GROUP BY 1 ORDER BY 1
""").to_dataframe()

# Comparativo por segmento (Excel apresenta valores apenas em 4 segmentos)
seg_excel = {
    1: ('Campeoes', 0, 0),
    2: ('Fieis', 0, 0),
    3: ('Fieis em potencial', 0, 0),
    4: ('Novos clientes', 0, 0),
    5: ('Promessas', 0, 0),
    6: ('Precisando de atencao', 0, 0),
    7: ('Quase dormentes', 0, 0),
    8: ('Nao pode perder', 5, 23895.76),
    9: ('Em risco', 31, 96340.49),
    10: ('Hibernando', 21, 23562.42),
    11: ('Perdidos', 191, 231345.42),
}
seg_sys = {int(r['seg_num']): (str(r['segmento']), int(r['clientes']), float(r['faturamento']))
           for _, r in df_seg_sys.iterrows()}

comp = []
for seg_num, (nome, exc_cli, exc_fat) in seg_excel.items():
    sys_nome, sys_cli, sys_fat = seg_sys.get(seg_num, (nome, 0, 0.0))
    comp.append([nome, exc_cli, sys_cli, sys_cli - exc_cli, exc_fat, sys_fat, sys_fat - exc_fat])
df_comp = pd.DataFrame(comp, columns=['Segmento', 'Excel Cli', 'Sis Cli', 'Delta Cli',
                                       'Excel Fat (R$)', 'Sis Fat (R$)', 'Delta Fat (R$)'])

# === EXCEL ===
wb = Workbook()
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FILL_PRIMARY = PatternFill('solid', fgColor='1E1882')
FILL_HEADER = PatternFill('solid', fgColor='4844C8')
FILL_WARN = PatternFill('solid', fgColor='FFE066')
FILL_DANGER = PatternFill('solid', fgColor='F8C9C9')
FILL_OK = PatternFill('solid', fgColor='C9F8D2')
FILL_LIGHT = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_hdr(cell, fill=FILL_HEADER):
    cell.font = ARIAL_HDR
    cell.fill = fill
    cell.alignment = CENTER
    cell.border = BORDER


def style_cell(cell, fmt=None):
    cell.font = ARIAL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


# Aba 1 — Resumo
ws = wb.active
ws.title = 'Resumo Executivo'
ws.merge_cells('A1:G1')
ws['A1'] = 'AUDITORIA RFV FARMACIAS - Excel (planilha) x Sistema (Dashboard Nevoni)'
ws['A1'].font = ARIAL_TITLE
ws['A1'].fill = FILL_PRIMARY
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 30

info = [
    ('Periodo auditado:', '01/04/2025 a 30/04/2026'),
    ('Familia RFV:', 'FARMACIAS'),
    ('Fonte Excel:', 'RFV Farmacias 01-04-2025 ate 30-04-2026 (1).xlsx'),
    ('Fonte Sistema:', 'silver_comercial.silver_com_rfv_score'),
    ('Gerado em:', datetime.now().strftime('%d/%m/%Y %H:%M')),
]
for i, (k, v) in enumerate(info, start=3):
    ws.cell(i, 1, k).font = ARIAL_BOLD
    ws.cell(i, 2, v).font = ARIAL

# Alerta inicial — planilha problema
ws.merge_cells('A9:G9')
ws['A9'] = 'ATENCAO: a planilha Farmacias apresenta varios problemas estruturais (vide aba "Achados Excel")'
ws['A9'].font = ARIAL_BOLD
ws['A9'].fill = FILL_WARN
ws['A9'].alignment = CENTER

ws['A11'] = 'TOTAIS GERAIS'
ws['A11'].font = ARIAL_BOLD
hdrs = ['Metrica', 'Excel (planilha)', 'Sistema (dashboard)', 'Delta', 'Delta %']
for i, h in enumerate(hdrs):
    style_hdr(ws.cell(12, i + 1, h))

dados = [
    ('Clientes unicos', 248, 245, 245 - 248, (245 - 248) / 248),
    ('Faturamento (R$)', 375144.09, 437269.74, 437269.74 - 375144.09, (437269.74 - 375144.09) / 375144.09),
]
for ri, d in enumerate(dados, start=13):
    for ci, v in enumerate(d, start=1):
        c = ws.cell(ri, ci, v)
        style_cell(c)
        if ci in (2, 3, 4):
            c.number_format = '#,##0.00' if ri == 14 else '#,##0'
        if ci == 5:
            c.number_format = '0.0%'
            c.fill = FILL_OK if v > 0 else FILL_DANGER

# Comparativo por segmento
ws['A16'] = 'COMPARATIVO POR SEGMENTO RFV'
ws['A16'].font = ARIAL_BOLD
ws['A17'] = 'Nota: Excel tem ZERO em 7 dos 11 segmentos. Sistema distribui mais granularmente. Investigar erros de formula no Excel.'
ws['A17'].font = ARIAL
ws['A17'].fill = FILL_WARN
ws.merge_cells('A17:G17')

for i, h in enumerate(df_comp.columns):
    style_hdr(ws.cell(18, i + 1, h))
for ri, row in enumerate(df_comp.itertuples(index=False), start=19):
    for ci, v in enumerate(row, start=1):
        c = ws.cell(ri, ci, v)
        style_cell(c)
        if ci > 1:
            c.alignment = RIGHT
        if ci in (5, 6, 7):
            c.number_format = '#,##0.00'
        if ci == 4:  # delta clientes
            if v > 0:
                c.fill = FILL_OK
            elif v < 0:
                c.fill = FILL_DANGER
        if ci == 7:  # delta fat
            if v > 0:
                c.fill = FILL_OK
            elif v < 0:
                c.fill = FILL_DANGER

# Pauta
ws['A32'] = 'PONTOS PARA REUNIAO COM ALVES/DIEGO'
ws['A32'].font = ARIAL_BOLD
pauta = [
    '1. Planilha de FARMACIAS apresenta problemas estruturais:',
    '   - Diversos #N/A nas formulas da matriz "Resultado RFV"',
    '   - Segunda regua de recencia no rodape (90/120/180/240/360) inconsistente com a regua superior',
    '   - 7 dos 11 segmentos com ZERO clientes - sugere erro de formula, nao realidade de negocio',
    '   - Base inicial contem duplicatas evidentes (linhas repetidas com mesmo CLIENTE/DATA/VALOR)',
    '',
    '2. Compare numerico Excel x Sistema:',
    '   - Volume total muito proximo: Excel 248 / Sistema 245 (3 a menos)',
    '   - Faturamento: Sistema R$ 437.270 vs Excel R$ 375.144 (sistema tem R$ 62k a mais)',
    '   - Sistema apresenta distribuicao por segmento que faz mais sentido de negocio',
    '',
    '3. Decisoes necessarias:',
    '   - Confiamos no Sistema como referencia para FARMACIAS (entendendo que a planilha esta com bugs)?',
    '   - Ou precisa-se reconstruir a planilha de FARMACIAS antes de comparar?',
    '   - Validar 3 clientes ausentes do sistema: ALESSANDRO ANTONIO DE MORAIS EPP, FARMACIA MAR CORREAS, FARMACIA SIMOES',
    '',
    '4. AUTOMACAO (continuacao da pauta Hospitalar):',
    '   - Reforca-se a necessidade de eliminar dependencia de planilhas manuais',
    '   - Para Farmacias, sugerir nova rodada do populate_carteira com fonte automatica',
]
for i, line in enumerate(pauta, start=33):
    c = ws.cell(i, 1, line)
    c.font = ARIAL
    if line.startswith(('1.', '2.', '3.', '4.')):
        c.font = ARIAL_BOLD

ws.column_dimensions['A'].width = 60
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col_letter].width = 16

# Aba 2 — Achados Excel
ws2 = wb.create_sheet('Achados Excel')
ws2.merge_cells('A1:D1')
ws2['A1'] = 'PROBLEMAS IDENTIFICADOS NA PLANILHA DE FARMACIAS'
ws2['A1'].font = ARIAL_TITLE
ws2['A1'].fill = FILL_PRIMARY
ws2['A1'].alignment = CENTER
ws2.row_dimensions[1].height = 30

hdrs = ['#', 'Achado', 'Evidencia', 'Impacto']
for i, h in enumerate(hdrs):
    style_hdr(ws2.cell(3, i + 1, h))

achados = [
    ('1', 'Celulas #N/A na matriz Resultado RFV',
     'Aba "Resultado RFV" linhas 4, 5, 8, 9, 10, 13, 14: contagens e valores #N/A',
     'Matriz inutilizavel - 7 dos 11 segmentos com 0 clientes'),
    ('2', 'Duas reguas de recencia diferentes',
     'Linha 1-2: R1=30d, R2=60d, R3=120d. Linha 15: 90d/120d/180d/240d/360d',
     'Inconsistencia interna - qual e a regua correta para Farmacias?'),
    ('3', 'Duplicatas na base inicial',
     f'{len(df_dups)} linhas suspeitas de duplicacao (CLIENTE+DATA+VALOR identicos). Ex: DROGARIA CELIA 29/04 R$ 940,53 aparece 2x',
     'Inflacionamento do count de pedidos por cliente'),
    ('4', 'Aba "Remover duplicatas" tem 1000 linhas (vazia/template?)',
     'Aba existe mas parece nao processada',
     'Sugere que o dedup nao foi executado nessa rodada'),
    ('5', 'Total Geral Perdidos = 191 (77% da base)',
     'Linha 22 da matriz: 191 clientes em Perdidos',
     'Distribuicao improvavel - sugere problema de calculo, nao realidade'),
]
for ri, ach in enumerate(achados, start=4):
    for ci, v in enumerate(ach, start=1):
        c = ws2.cell(ri, ci, v)
        c.font = ARIAL
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical='top',
                                horizontal='center' if ci == 1 else 'left')
    ws2.row_dimensions[ri].height = 50

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 38
ws2.column_dimensions['C'].width = 55
ws2.column_dimensions['D'].width = 38

# Aba 3 — 3 clientes faltantes + drill
ws3 = wb.create_sheet('Clientes Faltantes')
ws3.merge_cells('A1:D1')
ws3['A1'] = '3 CLIENTES PRESENTES NO EXCEL MAS AUSENTES DO SISTEMA - R$ 1.724,98'
ws3['A1'].font = ARIAL_TITLE
ws3['A1'].fill = FILL_PRIMARY
ws3['A1'].alignment = CENTER
ws3.row_dimensions[1].height = 30

hdrs = ['Cliente (Excel)', 'Pedidos', 'Valor (R$)', 'Ultima compra']
for i, h in enumerate(hdrs):
    style_hdr(ws3.cell(3, i + 1, h))

for ri, row in enumerate(so_excel.sort_values('valor', ascending=False).itertuples(index=False), start=4):
    cliente, pedidos, valor, ultima, k = row
    ws3.cell(ri, 1, cliente)
    ws3.cell(ri, 2, int(pedidos))
    ws3.cell(ri, 3, valor)
    ws3.cell(ri, 3).number_format = '#,##0.00'
    ws3.cell(ri, 4, ultima)
    ws3.cell(ri, 4).number_format = 'dd/mm/yyyy'
    for ci in range(1, 5):
        ws3.cell(ri, ci).font = ARIAL
        ws3.cell(ri, ci).border = BORDER
        if ci > 1:
            ws3.cell(ri, ci).alignment = RIGHT

ws3.column_dimensions['A'].width = 45
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 14

# Aba 4 — Drill clientes em comum
ws4 = wb.create_sheet('Drill 245 em comum')
ws4.merge_cells('A1:G1')
ws4['A1'] = 'COMPARATIVO PEDIDOS x VALOR - 245 CLIENTES EM COMUM (Sistema tem R$ 63k a mais)'
ws4['A1'].font = ARIAL_TITLE
ws4['A1'].fill = FILL_PRIMARY
ws4['A1'].alignment = CENTER
ws4.row_dimensions[1].height = 30

hdrs = ['Cliente', 'Excel Pedidos', 'Sis Pedidos', 'Excel Valor (R$)', 'Sis Valor (R$)',
        'Delta Pedidos', 'Delta Valor (R$)']
for i, h in enumerate(hdrs):
    style_hdr(ws4.cell(3, i + 1, h))

m_sorted = m.sort_values('delta_valor', ascending=False)
for ri, row in enumerate(m_sorted.itertuples(index=False), start=4):
    cliente, ped_b, val, ult, k, val_s, ped_s, dv, dp = row
    ws4.cell(ri, 1, cliente)
    ws4.cell(ri, 2, int(ped_b))
    ws4.cell(ri, 3, int(ped_s))
    ws4.cell(ri, 4, val)
    ws4.cell(ri, 5, val_s)
    ws4.cell(ri, 6, int(dp))
    ws4.cell(ri, 7, dv)
    for ci in range(1, 8):
        ws4.cell(ri, ci).font = ARIAL
        ws4.cell(ri, ci).border = BORDER
        if ci > 1:
            ws4.cell(ri, ci).alignment = RIGHT
        if ci in (4, 5, 7):
            ws4.cell(ri, ci).number_format = '#,##0.00'
    if dv > 100:
        ws4.cell(ri, 7).fill = FILL_OK
    elif dv < -100:
        ws4.cell(ri, 7).fill = FILL_DANGER
    if dp < 0:
        ws4.cell(ri, 6).fill = FILL_DANGER  # sistema tem menos pedidos (potencial duplicata Excel)
    elif dp > 0:
        ws4.cell(ri, 6).fill = FILL_OK

ws4.column_dimensions['A'].width = 55
for col_letter in 'BCDEFG':
    ws4.column_dimensions[col_letter].width = 15

# Aba 5 — Duplicatas detectadas
ws5 = wb.create_sheet('Duplicatas no Excel')
ws5.merge_cells('A1:D1')
ws5['A1'] = f'POTENCIAIS DUPLICATAS NA BASE INICIAL DO EXCEL ({len(df_dups)} linhas)'
ws5['A1'].font = ARIAL_TITLE
ws5['A1'].fill = FILL_PRIMARY
ws5['A1'].alignment = CENTER
ws5.row_dimensions[1].height = 30

hdrs = ['Cliente', 'Data', 'Valor (R$)', 'Observacao']
for i, h in enumerate(hdrs):
    style_hdr(ws5.cell(3, i + 1, h))

for ri, row in enumerate(df_dups.head(100).itertuples(index=False), start=4):
    cliente, data, valor, dup = row
    ws5.cell(ri, 1, cliente)
    ws5.cell(ri, 2, data)
    ws5.cell(ri, 2).number_format = 'dd/mm/yyyy'
    ws5.cell(ri, 3, valor)
    ws5.cell(ri, 3).number_format = '#,##0.00'
    ws5.cell(ri, 4, 'Linha repetida (mesmo cliente+data+valor)')
    for ci in range(1, 5):
        ws5.cell(ri, ci).font = ARIAL
        ws5.cell(ri, ci).border = BORDER

ws5.column_dimensions['A'].width = 50
ws5.column_dimensions['B'].width = 14
ws5.column_dimensions['C'].width = 14
ws5.column_dimensions['D'].width = 45

out = r'C:\Users\gusta\Downloads\Auditoria_RFV_Farmacias_Abril2026.xlsx'
wb.save(out)
print(f'OK: {out}')
