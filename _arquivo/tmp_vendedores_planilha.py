"""Lista exaustivamente quais vendedores aparecem nas 3 planilhas Alves."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl import load_workbook

pd.set_option('display.width', 200)
pd.set_option('display.max_colwidth', 80)

files = {
    'HOSPITALAR': r'C:\Users\gusta\Downloads\RFV Hospitalar 01-04-2025 até 30-04-2026 (1).xlsx',
    'FARMACIAS':  r'C:\Users\gusta\Downloads\RFV Farmácias 01-04-2025 até 30-04-2026 (1).xlsx',
    'SAC':        r'C:\Users\gusta\Downloads\RFV SAC 01-04-2025 até 30-04-2026 (1).xlsx',
}

NOMES_VENDEDORES = [
    'GUILHERME','KAUA','KAUÃ','RICHARD','GIOVAN','GEOVAN','EDUARDO','ALVES','VICTOR',
    'KARINA','CLARICE','CAUA','CAUÃ','RAMOS','GABRIEL','RIBEIRO','EDGARD','LUAN',
    'ABNER','AUGUSTO','MESQUITA','VINICIUS','VINÍCIUS','SOARES','OLIVEIRA',
    'MARQUES','RODRIGUES','SEQUIM','SABINO','LUCAS','BERNARDINO',
]

for fam, path in files.items():
    print('=' * 90)
    print(f'{fam}')
    print('=' * 90)
    xls = pd.ExcelFile(path)
    print(f'\nAbas:')
    for s in xls.sheet_names:
        print(f'  - "{s}"')

    print(f'\n>>> Procurando nomes de vendedores em CADA aba (cells + columns + sheet names):')

    # 1) Search in sheet names
    for sheet in xls.sheet_names:
        sheet_upper = sheet.upper()
        for v in NOMES_VENDEDORES:
            if v in sheet_upper:
                print(f'  [SHEET NAME] "{sheet}" contém "{v}"')

    # 2) Search header cells (linha de cabeçalho de cada aba)
    print(f'\n>>> Headers (linha 1-3) de cada aba que contém nome de vendedor:')
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in xls.sheet_names:
        try:
            ws = wb[sheet]
            for row in list(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
                for cell in row:
                    if cell is None:
                        continue
                    cv = str(cell).upper()
                    for v in NOMES_VENDEDORES:
                        if v in cv and v not in cv.replace('GIOVANI', ''):  # exclui falsos positivos cliente
                            if 'LTDA' not in cv and 'RODRIGUES' not in cv.replace('KAUA RODRIGUES','').replace('KAUÃ RODRIGUES','') \
                               or v in ['KAUA','KAUÃ','RIBEIRO','RAMOS']:
                                print(f'  [{sheet} | HEADER] "{cell}"')
                                break
        except Exception as e:
            pass

    # 3) Procura especifica por palavras-chave de classificacao (vendedor, responsavel)
    print(f'\n>>> Aba "Apoio" e similares (provavelmente o mapa A/B/C → vendedor):')
    for sheet in xls.sheet_names:
        if 'APOIO' in sheet.upper() or 'AUX' in sheet.upper() or 'LEGEND' in sheet.upper():
            try:
                df = pd.read_excel(path, sheet_name=sheet, header=None)
                print(f'\n  Aba "{sheet}":')
                print(df.head(30).to_string(index=False, header=False))
            except Exception as e:
                print(f'  Erro: {e}')

print()
print('=' * 90)
print('NOTA: As abas de vendedor "Base inicial - Vendedor A/B/C" não dizem')
print('o nome real do vendedor — A/B/C são códigos. O nome real fica em:')
print('  - Aba "Apoio" (se tiver)')
print('  - OU foi passado verbalmente pelo Alves e codificado em populate_carteira.py')
print('=' * 90)
