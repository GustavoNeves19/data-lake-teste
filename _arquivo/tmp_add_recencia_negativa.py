"""Adiciona aba 'Recencias Negativas' nas auditorias do Hospitalar."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
ARIAL_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ARIAL_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
ARIAL_SUB = Font(name='Arial', size=11, bold=True)
FILL_PRIMARY = PatternFill('solid', fgColor='1E1882')
FILL_HEADER = PatternFill('solid', fgColor='4844C8')
FILL_WARN = PatternFill('solid', fgColor='FFE066')
FILL_DANGER = PatternFill('solid', fgColor='F8C9C9')
FILL_OK = PatternFill('solid', fgColor='C9F8D2')
FILL_LIGHT = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Carrega dados
path_excel = r'C:\Users\gusta\Downloads\RFV Hospitalar 01-04-2025 até 30-04-2026 (1).xlsx'
df_bi = pd.read_excel(path_excel, sheet_name='Base inicial - Geral', usecols=[0,1,2])
df_bi.columns = ['cliente', 'data_venda', 'valor']
df_bi = df_bi.dropna(subset=['cliente'])
df_bi['data_venda'] = pd.to_datetime(df_bi['data_venda'])
df_bi['valor'] = pd.to_numeric(df_bi['valor'])
df_bi['k'] = df_bi['cliente'].str.upper().str.strip()

df_sf = pd.read_excel(path_excel, sheet_name='Sem fórmula Geral')
df_sf['k'] = df_sf['ID - CLIENTE'].astype(str).str.upper().str.strip()
df_sf['Data última compra'] = pd.to_datetime(df_sf['Data última compra'])

ref_planilha = pd.Timestamp('2026-04-02')

# Clientes com compra apos 02/04
apos = df_bi[df_bi['data_venda'] > ref_planilha]
clientes_apos = apos.groupby('k').agg(
    cliente=('cliente', 'first'),
    pedidos_apos=('data_venda', 'count'),
    valor_apos=('valor', 'sum'),
    ultima_real=('data_venda', 'max'),
    primeira_apos=('data_venda', 'min'),
).reset_index()

# Merge com Sem Formula
cmp = clientes_apos.merge(
    df_sf[['k','Data última compra','Recência em dias','Recência 3','Frequência 1','Frequência 2','Classificação 2','Valor']],
    on='k', how='left'
)
cmp = cmp.sort_values('valor_apos', ascending=False)

# Total
total_clientes = len(cmp)
total_pedidos = int(cmp['pedidos_apos'].sum())
total_valor = float(cmp['valor_apos'].sum())
total_neg = int((cmp['Recência em dias'] < 0).sum())

print(f'Clientes com compra pos-02/04: {total_clientes}')
print(f'Pedidos pos-02/04: {total_pedidos}')
print(f'Valor pos-02/04: R$ {total_valor:,.2f}')
print(f'Com recencia NEGATIVA na planilha: {total_neg}')


def adiciona_aba(wb_path):
    wb = load_workbook(wb_path)
    nome_aba = 'Recencias Negativas'
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    # Posicionar logo apos Resumo Executivo
    pos = 1
    if 'Resumo Executivo' in wb.sheetnames:
        pos = wb.sheetnames.index('Resumo Executivo') + 1
    ws = wb.create_sheet(nome_aba, pos)

    # Titulo
    ws.merge_cells('A1:H1')
    ws['A1'] = 'BUG OCULTO DA PLANILHA HOSPITALAR — RECENCIA NEGATIVA'
    ws['A1'].font = ARIAL_TITLE
    ws['A1'].fill = FILL_PRIMARY
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Contexto
    ws.merge_cells('A2:H2')
    ws['A2'] = ('A planilha promete dados ate 30/04/2026 (vide nome do arquivo) e CUMPRE — a Base Inicial tem '
                'compras ate 30/04. Porem, a celula "Data de hoje" foi preenchida em 02/04/2026 e nunca atualizada. '
                'Resultado: compras posteriores a 02/04 geram recencia NEGATIVA na planilha.')
    ws['A2'].font = ARIAL
    ws['A2'].fill = FILL_WARN
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 50

    # Numeros agregados
    ws['A4'] = 'NUMEROS DO BUG'
    ws['A4'].font = ARIAL_SUB
    hdrs = ['Metrica', 'Valor']
    for i, h in enumerate(hdrs):
        c = ws.cell(5, i + 1, h)
        c.font = ARIAL_HDR; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER

    stats = [
        ('Periodo declarado no nome do arquivo', '01/04/2025 a 30/04/2026'),
        ('Data de "hoje" usada nas formulas', '02/04/2026'),
        ('Gap (dias ignorados pelo calculo)', '28 dias'),
        ('Linhas de compra com data > 02/04/2026 na Base Inicial', total_pedidos),
        ('Clientes afetados (com pelo menos 1 compra apos 02/04)', total_clientes),
        ('Valor total dessas compras ignoradas pelo calculo correto', f'R$ {total_valor:,.2f}'),
        ('Clientes com RECENCIA NEGATIVA na "Sem formula"', total_neg),
    ]
    for ri, (k, v) in enumerate(stats, start=6):
        ws.cell(ri, 1, k).font = ARIAL
        ws.cell(ri, 1).border = BORDER
        ws.cell(ri, 2, v).font = ARIAL
        ws.cell(ri, 2).alignment = RIGHT
        ws.cell(ri, 2).border = BORDER
        if 'NEGATIVA' in k or 'RECENCIA' in k:
            ws.cell(ri, 1).fill = FILL_DANGER
            ws.cell(ri, 2).fill = FILL_DANGER

    # Por que passou despercebido
    ws['A15'] = 'POR QUE O BUG PASSOU DESPERCEBIDO'
    ws['A15'].font = ARIAL_SUB
    explic = [
        '1. Recencia NEGATIVA cai automaticamente em R1 (regra: recencia <= 30 dias).',
        '2. Esses clientes compraram em abril/2026 — DEVERIAM mesmo estar em R1 (recencia real de 0 a 28 dias positivos).',
        '3. O erro entao "se compensa" — a classificacao acaba acidentalmente correta por borda de bucket.',
        '4. PORÉM, conceitualmente a planilha esta errada. A "Data de hoje" deveria ser 30/04/2026 (fim do periodo).',
        '5. Em familias com regua de R diferente (Farmacias), esse mesmo bug pode produzir classificacoes erradas.',
    ]
    for i, line in enumerate(explic, start=16):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        c = ws.cell(i, 1, line)
        c.font = ARIAL
        c.alignment = Alignment(wrap_text=True, vertical='center')

    # Tabela: clientes com recencia negativa
    titulo_row = 23
    ws.merge_cells(start_row=titulo_row, start_column=1, end_row=titulo_row, end_column=8)
    ws.cell(titulo_row, 1, f'OS {total_clientes} CLIENTES COM COMPRA POS-02/04 — TOP 30 POR VALOR')
    ws.cell(titulo_row, 1).font = ARIAL_SUB

    hdr_row = titulo_row + 1
    hdrs = ['Cliente', 'Ultima compra REAL', 'Pedidos pos-02/04', 'Valor pos-02/04 (R$)',
            'Ultima na planilha', 'Recencia planilha (dias)', 'R-bucket', 'Segmento planilha']
    for i, h in enumerate(hdrs):
        c = ws.cell(hdr_row, i + 1, h)
        c.font = ARIAL_HDR; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER

    for ri, r in enumerate(cmp.head(30).itertuples(index=False), start=hdr_row + 1):
        k, cliente, ped_apos, val_apos, ult_real, prim_apos, ult_pl, rec_pl, rec_b, freq1, freq2, seg, val = r
        ws.cell(ri, 1, cliente)
        ws.cell(ri, 2, ult_real)
        ws.cell(ri, 2).number_format = 'dd/mm/yyyy'
        ws.cell(ri, 3, int(ped_apos))
        ws.cell(ri, 4, float(val_apos))
        ws.cell(ri, 4).number_format = '#,##0.00'
        if pd.notna(ult_pl):
            ws.cell(ri, 5, ult_pl)
            ws.cell(ri, 5).number_format = 'dd/mm/yyyy'
        if pd.notna(rec_pl):
            rec_int = int(rec_pl)
            ws.cell(ri, 6, rec_int)
            if rec_int < 0:
                ws.cell(ri, 6).fill = FILL_DANGER
                ws.cell(ri, 6).font = ARIAL_BOLD
        ws.cell(ri, 7, rec_b)
        ws.cell(ri, 8, seg)
        for ci in range(1, 9):
            cell = ws.cell(ri, ci)
            if cell.font.name != 'Arial' or not cell.font.bold:
                cell.font = ARIAL
            cell.border = BORDER
            if ci in (3, 4, 6):
                cell.alignment = RIGHT

    # Conclusao
    fim_row = hdr_row + 32
    ws.merge_cells(start_row=fim_row, start_column=1, end_row=fim_row, end_column=8)
    c = ws.cell(fim_row, 1,
        'CONCLUSAO: ate a planilha "mais integra" das tres tem inconsistencia entre dados coletados '
        '(ate 30/04) e referencia temporal usada nas formulas (02/04). Argumento adicional pela automacao da carteira.')
    c.font = ARIAL_BOLD
    c.fill = FILL_OK
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[fim_row].height = 35

    widths = [55, 14, 12, 16, 14, 14, 9, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    wb.save(wb_path)


# Atualiza os dois arquivos do Hospitalar
adiciona_aba(r'C:\Users\gusta\Downloads\Auditoria_RFV_Hospitalar_Abril2026.xlsx')
adiciona_aba(r'C:\Users\gusta\Downloads\Auditoria_RFV_Hospitalar_Classificacoes.xlsx')
print('OK: Aba "Recencias Negativas" adicionada nos dois arquivos do Hospitalar.')
